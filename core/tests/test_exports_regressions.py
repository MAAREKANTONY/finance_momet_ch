from django.test import SimpleTestCase, TestCase

from core.views import _arrow_table_to_csv_safe


class ExportRegressionTests(SimpleTestCase):
    def test_arrow_csv_safe_helper_is_plain_function(self):
        class DummyTable:
            column_names = []

        table = DummyTable()
        result = _arrow_table_to_csv_safe(table)
        self.assertIs(result, table)


from openpyxl import load_workbook
from tempfile import NamedTemporaryFile, TemporaryDirectory
from types import SimpleNamespace
import gzip
import json

from core.backtest_debug import build_backtest_debug_workbook
from core.models import Backtest
from core.views import _build_backtest_workbook_full


class ExcelSerializationRegressionTests(SimpleTestCase):
    def _sheet_flat_rows(self, workbook, sheet_name):
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        return [" | ".join("" if cell is None else str(cell) for cell in row) for row in rows]

    def _make_backtest_stub(self):
        scenario = SimpleNamespace(
            name="Scenario X",
            description="",
            a=1, b=1, c=1, d=1, e=0, vc=None, fl=None, n1=1, n2=1, n3=1, n4=1,
            n5=1, k2j=None, cr=None, n5f3=None, crf3=None,
            npente=None, nglobal=None,
            slope_threshold=None, slope_sell_threshold=None,
            npente_basse=None, slope_threshold_basse=None, slope_sell_threshold_basse=None,
            recent_high_drawdown_lookback_days=None, recent_high_drawdown_max_drop_pct=None,
        )
        results = {
            "tickers": {
                "AAA": {
                    "lines": [
                        {
                            "line_index": 1,
                            "buy": ["SPA", "SPVA"],
                            "sell": ["SVA"],
                            "allocated": True,
                            "final": {"N": 2, "S_G_N": 0.1, "BT": 0.2, "NB_JOUR_OUVRES": 3, "BMJ": 0.01, "BMD": 0.02, "BUY_DAYS_CLOSED": 1, "cash_ticker_end": 123.4},
                            "daily": [],
                        }
                    ]
                }
            },
            "portfolio": {"kpi": {}, "daily": []},
            "meta": {},
        }
        return SimpleNamespace(id=1, name="BT", description="", scenario=scenario, scenario_id=1, start_date=None, end_date=None, capital_total=0, capital_per_ticker=0, ratio_threshold=0, close_positions_at_end=False, status="DONE", universe_snapshot=[], results=results, capital_mode="fixed", include_all_tickers=False, warmup_days=0)

    def test_build_backtest_workbook_compact_serializes_list_cells(self):
        bt = self._make_backtest_stub()
        bt.results["tickers"]["AAA"]["lines"][0]["buy_market_gm_market"] = "GM_POS"
        from core.views import _build_backtest_workbook_compact
        wb, _ = _build_backtest_workbook_compact(bt)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["Summary"].iter_rows(values_only=True))
        self.assertEqual(rows[1][2], '["SPA", "SPVA"]')
        self.assertEqual(rows[1][3], 'GM marché: GM positif')
        self.assertEqual(rows[1][4], "Aucune")
        self.assertEqual(rows[1][5], "Aucune")
        self.assertEqual(rows[1][6], "Aucune")
        self.assertEqual(rows[1][7], '["SVA"]')

    def test_build_backtest_workbook_full_serializes_list_cells(self):
        bt = self._make_backtest_stub()
        bt.results["tickers"]["AAA"]["lines"][0]["buy_market_gm_market"] = "GM_POS"
        wb, _ = _build_backtest_workbook_full(bt)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["Summary"].iter_rows(values_only=True))
        self.assertEqual(rows[1][2], '["SPA", "SPVA"]')
        self.assertEqual(rows[1][3], 'GM marché: GM positif')
        self.assertEqual(rows[1][4], "Aucune")
        self.assertEqual(rows[1][5], "Aucune")
        self.assertEqual(rows[1][6], "Aucune")
        self.assertEqual(rows[1][7], '["SVA"]')

    def test_build_backtest_workbook_full_settings_include_dynamic_universe_metadata(self):
        bt = self._make_backtest_stub()
        bt.results["meta"]["universe"] = {
            "mode": "SP500_HISTORICAL_DYNAMIC",
            "universe_code": "SP500",
            "universe_name": "S&P 500",
            "coverage_start": "2020-01-01",
            "coverage_end": "2020-12-31",
            "superset_count": 503,
            "source": "manual_csv",
        }
        wb, _ = _build_backtest_workbook_full(bt)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            flat = self._sheet_flat_rows(loaded, "Settings")
        self.assertTrue(any("Univers mode | SP500_HISTORICAL_DYNAMIC" in row for row in flat))
        self.assertTrue(any("Univers code | SP500" in row for row in flat))
        self.assertTrue(any("Univers nom | S&P 500" in row for row in flat))
        self.assertTrue(any("Univers période historique début | 2020-01-01" in row for row in flat))
        self.assertTrue(any("Univers période historique fin | 2020-12-31" in row for row in flat))
        self.assertTrue(any("Univers actions analysées | 503" in row for row in flat))
        self.assertTrue(any("Univers source des données | manual_csv" in row for row in flat))

    def test_build_backtest_workbook_compact_settings_include_dynamic_universe_metadata(self):
        bt = self._make_backtest_stub()
        bt.results["meta"]["universe"] = {
            "mode": "SP500_HISTORICAL_DYNAMIC",
            "universe_code": "SP500",
            "coverage_start": "2020-01-01",
            "coverage_end": "2020-12-31",
            "superset_count": 503,
            "source": "manual_csv",
        }
        from core.views import _build_backtest_workbook_compact

        wb, _ = _build_backtest_workbook_compact(bt, charts="0")
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            flat = self._sheet_flat_rows(loaded, "Settings")
        self.assertTrue(any("Univers mode | SP500_HISTORICAL_DYNAMIC" in row for row in flat))
        self.assertTrue(any("Univers code | SP500" in row for row in flat))
        self.assertTrue(any("Univers période historique début | 2020-01-01" in row for row in flat))
        self.assertTrue(any("Univers période historique fin | 2020-12-31" in row for row in flat))
        self.assertTrue(any("Univers actions analysées | 503" in row for row in flat))
        self.assertTrue(any("Univers source des données | manual_csv" in row for row in flat))

    def test_backtest_exports_include_csi300_effective_currency(self):
        bt = self._make_backtest_stub()
        bt.scenario.universe_mode = "CSI300_HISTORICAL_DYNAMIC"
        bt.results["meta"]["effective_currency"] = "CNY"

        full, _ = _build_backtest_workbook_full(bt)
        from core.views import _build_backtest_workbook_compact
        compact, _ = _build_backtest_workbook_compact(bt, charts="0")

        for workbook in (full, compact):
            flat = self._sheet_flat_rows(workbook, "Settings")
            self.assertTrue(any("Devise effective | CNY" in row for row in flat))

    def test_old_static_backtest_exports_ignore_mutable_csi300_scenario_and_settings(self):
        bt = self._make_backtest_stub()
        bt.scenario.universe_mode = "CSI300_HISTORICAL_DYNAMIC"
        bt.settings = {"effective_currency": "CNY"}

        full, _ = _build_backtest_workbook_full(bt)
        from core.views import _build_backtest_workbook_compact
        compact, _ = _build_backtest_workbook_compact(bt, charts="0")

        for workbook in (full, compact):
            flat = self._sheet_flat_rows(workbook, "Settings")
            self.assertFalse(any("Devise effective" in row for row in flat))

    def test_done_backtest_with_empty_results_does_not_export_settings_currency(self):
        bt = self._make_backtest_stub()
        bt.status = Backtest.Status.DONE
        bt.results = {}
        bt.settings = {"effective_currency": "CNY"}
        original_results = dict(bt.results)
        original_settings = dict(bt.settings)

        from core.views import _build_backtest_workbook_compact

        with self.assertRaisesMessage(ValueError, "Aucun résultat à exporter"):
            _build_backtest_workbook_full(bt)
        with self.assertRaisesMessage(ValueError, "Aucun résultat à exporter"):
            _build_backtest_workbook_compact(bt, charts="0")
        self.assertEqual(bt.results, original_results)
        self.assertEqual(bt.settings, original_settings)

    def test_build_backtest_workbook_full_uses_bounded_return_wording_for_global_momentum(self):
        bt = self._make_backtest_stub()
        bt.results["portfolio"]["daily"] = [
            {
                "date": "2024-01-02",
                "equity": "1000",
                "invested": "900",
                "global_cash": "100",
                "cash_allocated": "0",
                "positions_value": "900",
                "pnl_global": "100",
                "portfolio_return_global": "0.1",
                "avg_global_nglobal": "5",
                "drawdown": "0",
            }
        ]
        wb, _ = _build_backtest_workbook_full(bt)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["Portfolio_Daily"].iter_rows(values_only=True))
        self.assertIn("Moyenne globale rendements bornés Nglobal (%)", rows[0][8])

    def test_build_backtest_workbook_full_summary_includes_buy_and_sell_slope_thresholds(self):
        bt = self._make_backtest_stub()
        bt.scenario.npente = 100
        bt.scenario.slope_threshold = "0.10"
        bt.scenario.slope_sell_threshold = "0.05"
        bt.scenario.npente_basse = 20
        bt.scenario.slope_threshold_basse = "0.02"
        bt.scenario.slope_sell_threshold_basse = "0.01"
        wb, _ = _build_backtest_workbook_full(bt)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["Settings"].iter_rows(values_only=True))
        flat = [" | ".join("" if cell is None else str(cell) for cell in row) for row in rows]
        self.assertTrue(any("SUM_SLOPE seuil achat | 0.10" in row for row in flat))
        self.assertTrue(any("SUM_SLOPE seuil vente | 0.05" in row for row in flat))
        self.assertTrue(any("SUM_SLOPE_BASSE seuil achat | 0.02" in row for row in flat))
        self.assertTrue(any("SUM_SLOPE_BASSE seuil vente | 0.01" in row for row in flat))

    def test_build_backtest_workbook_full_summary_includes_recent_high_drawdown_settings(self):
        bt = self._make_backtest_stub()
        bt.scenario.recent_high_drawdown_lookback_days = 10
        bt.scenario.recent_high_drawdown_max_drop_pct = "-0.10"
        wb, _ = _build_backtest_workbook_full(bt)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["Settings"].iter_rows(values_only=True))
        flat = [" | ".join("" if cell is None else str(cell) for cell in row) for row in rows]
        self.assertTrue(any("Signal anti-chute RHD fenêtre | 10" in row for row in flat))
        self.assertTrue(any("Signal anti-chute RHD repli max | -0.10" in row for row in flat))

    def test_backtest_debug_workbook_serializes_nested_daily_values(self):
        scenario = SimpleNamespace(name="Scenario X", description="")
        bt = SimpleNamespace(id=1, name="BT", scenario=scenario, start_date=None, end_date=None, capital_total=0, capital_per_ticker=0, capital_mode="fixed", ratio_threshold=0, include_all_tickers=False, warmup_days=0, close_positions_at_end=False, results={
            "tickers": {
                "AAA": {
                    "lines": [{
                        "line_index": 1,
                        "buy": ["SPA", "SPVA"],
                        "sell": ["SVA"],
                        "daily": [{"date": "2026-01-01", "action": ["BUY", "SELL"]}],
                        "final": {"alerts": ["A", "B"]},
                    }]
                }
            }
        })
        wb, _ = build_backtest_debug_workbook(bt, ticker="AAA", line=1)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["DATA"].iter_rows(values_only=True))
        self.assertIn('["BUY", "SELL"]', rows[1])

    def test_backtest_debug_workbook_exposes_dynamic_universe_daily_fields(self):
        scenario = SimpleNamespace(name="Scenario X", description="")
        bt = SimpleNamespace(id=1, name="BT", scenario=scenario, start_date=None, end_date=None, capital_total=0, capital_per_ticker=0, capital_mode="fixed", ratio_threshold=0, include_all_tickers=False, warmup_days=0, close_positions_at_end=False, results={
            "tickers": {
                "AAA": {
                    "lines": [{
                        "line_index": 1,
                        "buy": ["SPA"],
                        "sell": ["SVA"],
                        "daily": [{
                            "date": "2026-01-01",
                            "universe_member": False,
                            "buy_blocked_by_universe": True,
                            "buy_blocked_reason": "not_active_in_universe",
                        }],
                        "final": {},
                    }]
                }
            }
        })
        wb, _ = build_backtest_debug_workbook(bt, ticker="AAA", line=1)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["DATA"].iter_rows(values_only=True))
        headers = list(rows[0])
        values = dict(zip(headers, rows[1]))
        self.assertIn("universe_member", headers)
        self.assertIn("buy_blocked_by_universe", headers)
        self.assertIn("buy_blocked_reason", headers)
        self.assertIs(values["universe_member"], False)
        self.assertIs(values["buy_blocked_by_universe"], True)
        self.assertEqual(values["buy_blocked_reason"], "not_active_in_universe")

    def test_backtest_debug_workbook_handles_large_result_mode_without_daily_rows(self):
        scenario = SimpleNamespace(name="Scenario X", description="")
        bt = SimpleNamespace(id=1, name="BT", scenario=scenario, start_date=None, end_date=None, capital_total=0, capital_per_ticker=0, capital_mode="fixed", ratio_threshold=0, include_all_tickers=False, warmup_days=0, close_positions_at_end=False, results={
            "meta": {"large_result_mode": True, "detailed_daily_rows_omitted": True, "estimated_daily_rows": 900000},
            "tickers": {
                "AAA": {
                    "lines": [{
                        "line_index": 1,
                        "buy": ["SPA"],
                        "sell": ["SVA"],
                        "daily": [],
                        "daily_rows_omitted": True,
                        "final": {"N": 1, "BT": "0.2"},
                    }]
                }
            }
        })
        wb, _ = build_backtest_debug_workbook(bt, ticker="AAA", line=1)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["DATA"].iter_rows(values_only=True))
        self.assertEqual(rows[1][0], "Aucune ligne quotidienne disponible")

    def test_backtest_debug_workbook_lists_sell_threshold_fields_when_present(self):
        scenario = SimpleNamespace(
            name="Scenario X",
            description="",
            slope_threshold="0.10",
            slope_sell_threshold="0.05",
            npente_basse=20,
            slope_threshold_basse="0.02",
            slope_sell_threshold_basse="0.01",
        )
        bt = SimpleNamespace(id=1, name="BT", scenario=scenario, start_date=None, end_date=None, capital_total=0, capital_per_ticker=0, capital_mode="fixed", ratio_threshold=0, include_all_tickers=False, warmup_days=0, close_positions_at_end=False, results={
            "tickers": {
                "AAA": {
                    "lines": [{
                        "line_index": 1,
                        "buy": ["SPA"],
                        "sell": ["SPVv"],
                        "daily": [],
                        "final": {},
                    }]
                }
            }
        })
        wb, _ = build_backtest_debug_workbook(bt, ticker="AAA", line=1)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["FORMULAS"].iter_rows(values_only=True))
        flat = [" | ".join("" if cell is None else str(cell) for cell in row) for row in rows]
        self.assertTrue(any("Seuil de pente vente | 0.05" in row for row in flat))
        self.assertTrue(any("Seuil de pente basse vente | 0.01" in row for row in flat))

    def test_backtest_debug_workbook_lists_recent_high_drawdown_fields_when_present(self):
        scenario = SimpleNamespace(
            name="Scenario X",
            description="",
            recent_high_drawdown_lookback_days=10,
            recent_high_drawdown_max_drop_pct="-0.10",
        )
        bt = SimpleNamespace(id=1, name="BT", scenario=scenario, start_date=None, end_date=None, capital_total=0, capital_per_ticker=0, capital_mode="fixed", ratio_threshold=0, include_all_tickers=False, warmup_days=0, close_positions_at_end=False, results={
            "tickers": {
                "AAA": {
                    "lines": [{
                        "line_index": 1,
                        "buy": ["RHD_OK"],
                        "sell": ["RHD_FAIL"],
                        "daily": [],
                        "final": {},
                    }]
                }
            }
        })
        wb, _ = build_backtest_debug_workbook(bt, ticker="AAA", line=1)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["FORMULAS"].iter_rows(values_only=True))
        flat = [" | ".join("" if cell is None else str(cell) for cell in row) for row in rows]
        self.assertTrue(any("Signal anti-chute RHD fenêtre | 10" in row for row in flat))
        self.assertTrue(any("Signal anti-chute RHD repli max | -0.10" in row for row in flat))

    def test_backtest_debug_workbook_lists_line_market_conditions_when_present(self):
        scenario = SimpleNamespace(name="Scenario X", description="")
        bt = SimpleNamespace(id=1, name="BT", scenario=scenario, start_date=None, end_date=None, capital_total=0, capital_per_ticker=0, capital_mode="fixed", ratio_threshold=0, include_all_tickers=False, warmup_days=0, close_positions_at_end=False, results={
            "tickers": {
                "AAA": {
                    "lines": [{
                        "line_index": 1,
                        "buy": ["SPA"],
                        "buy_market_gm_current": "GM_POS",
                        "buy_market_gm_market": "GM_NEG",
                        "buy_market_operator": "AND",
                        "gm_sell_market_exit_conditions": {
                            "operator": "OR",
                            "market": {"mode": "GM_NEG", "threshold": "-0.03", "explicit_threshold": True},
                        },
                        "sell": ["SPVv"],
                        "daily": [],
                        "final": {},
                    }]
                }
            }
        })
        wb, _ = build_backtest_debug_workbook(bt, ticker="AAA", line=1)
        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            loaded = load_workbook(tmp.name, read_only=True)
            rows = list(loaded["FORMULAS"].iter_rows(values_only=True))
        flat = [" | ".join("" if cell is None else str(cell) for cell in row) for row in rows]
        self.assertTrue(any("Conditions de marché | GM actuel: GM positif ET GM marché: GM négatif" in row for row in flat))
        self.assertTrue(any("Protection marché GM | GM marché: GM négatif < -0.03" in row for row in flat))



class BacktestDebugCsvExportRegressionTests(TestCase):
    def test_backtest_debug_csv_reads_offloaded_daily_rows(self):
        from pathlib import Path
        from unittest.mock import patch
        from core.models import Backtest, ProcessingJob, Scenario
        from core.tasks import export_backtest_debug_csv_task

        scenario = Scenario.objects.create(name="Scenario CSV")
        with TemporaryDirectory() as tmp_dir:
            daily_path = Path(tmp_dir) / "aaa_L1.json.gz"
            with gzip.open(daily_path, "wt", encoding="utf-8") as handle:
                json.dump([
                    {
                        "date": "2026-01-01",
                        "universe_member": False,
                        "buy_blocked_by_universe": True,
                        "buy_blocked_reason": "not_active_in_universe",
                    }
                ], handle)
            bt = Backtest.objects.create(
                name="BT CSV",
                scenario=scenario,
                results={
                    "tickers": {
                        "AAA": {
                            "lines": [{
                                "line_index": 1,
                                "buy": ["SPA"],
                                "sell": ["SVA"],
                                "daily_offloaded": True,
                                "daily_backend": "json.gz",
                                "daily_path": str(daily_path),
                                "daily_rows": 1,
                                "final": {},
                            }]
                        }
                    }
                },
            )
            job = ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXPORT_BACKTEST_DEBUG_CSV, status=ProcessingJob.Status.PENDING)
            output_path = Path(tmp_dir) / "debug.csv"

            with patch("core.tasks._job_export_path", return_value=output_path):
                result = export_backtest_debug_csv_task.run(job_id=job.id, backtest_id=bt.id, ticker="AAA", line="1")

            self.assertEqual(result, str(output_path))
            content = output_path.read_text(encoding="utf-8")
        self.assertIn("universe_member", content)
        self.assertIn("buy_blocked_by_universe", content)
        self.assertIn("buy_blocked_reason", content)
        self.assertIn("not_active_in_universe", content)


class GameScenarioExportRegressionTests(TestCase):
    def test_export_game_scenario_xlsx_handles_dict_today_results(self):
        from unittest.mock import patch
        from core.models import GameScenario, ProcessingJob
        from core.tasks import export_game_scenario_xlsx_task

        game = GameScenario.objects.create(
            name='G1',
            today_results={'ticker': 'AF', 'ok': True},
            recent_high_drawdown_lookback_days=10,
            recent_high_drawdown_max_drop_pct='-0.10',
        )
        job = ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXPORT_GAME_SCENARIO_XLSX, status=ProcessingJob.Status.PENDING)

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            with patch('core.tasks._job_export_path', return_value=tmp.name):
                result = export_game_scenario_xlsx_task.run(job_id=job.id, game_scenario_id=game.id)

            loaded = load_workbook(tmp.name, read_only=True)
            game_rows = list(loaded["Game"].iter_rows(values_only=True))
        flat = [" | ".join("" if cell is None else str(cell) for cell in row) for row in game_rows]

        self.assertEqual(result, tmp.name)
        job.refresh_from_db()
        self.assertEqual(job.status, ProcessingJob.Status.DONE)
        self.assertEqual(job.output_file, tmp.name)
        self.assertTrue(any("Signal anti-chute RHD fenêtre | 10" in row for row in flat))
        self.assertTrue(any("Signal anti-chute RHD repli max | -0.1" in row or "Signal anti-chute RHD repli max | -0.10" in row for row in flat))
