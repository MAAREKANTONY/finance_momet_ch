from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from core.models import BacktestDailyStat, BacktestResult, BacktestRun, DailyMetric


def export_backtest_run_xlsx(run: BacktestRun) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Parameters
    ws.append(["Backtest name", run.name])
    ws.append(["Description", run.description])
    ws.append(["Scenario", run.scenario.name])
    ws.append(["Strategy", run.strategy.name])
    ws.append(["CP (capital total, 0=infinite)", float(run.capital_total)])
    ws.append(["CT (capital per symbol)", float(run.capital_per_symbol)])
    ws.append(["X (min ratio_p %)", float(run.min_ratio_p)])
    ws.append(["Created at", run.created_at.isoformat() if run.created_at else ""])
    ws.append(["Status", run.status])
    ws.append([])

    # Results table
    headers = ["Ticker", "Market", "Initial", "Final", "Return %", "#Trades", "Last close"]
    ws.append(headers)

    results = list(
        BacktestResult.objects.filter(run=run)
        .select_related("symbol")
        .order_by("symbol__ticker")
    )
    for r in results:
        ws.append([
            r.symbol.ticker,
            r.symbol.exchange,
            float(r.initial_capital) if r.initial_capital is not None else None,
            float(r.final_capital) if r.final_capital is not None else None,
            float(r.return_pct) if r.return_pct is not None else None,
            r.trades_count,
            float(r.last_close) if r.last_close is not None else None,
        ])

    # autosize
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # One sheet per symbol with daily stats and charts
    for r in results:
        sym = r.symbol
        sh = wb.create_sheet(title=sym.ticker[:31])

        sh.append(["Date", "ratio_p", "N", "G", "S_G_N", "BT", "TradableDays", "BMJ"])
        # Build ratio map for speed
        ratio_map = {m.date: m.RATIO_P for m in DailyMetric.objects.filter(symbol=sym, scenario=run.scenario).only("date", "RATIO_P")}
        stats = list(
            BacktestDailyStat.objects.filter(run=run, symbol=sym).order_by("date")
        )
        for st in stats:
            sh.append([
                st.date.isoformat(),
                float(ratio_map.get(st.date)) if ratio_map.get(st.date) is not None else None,
                st.N,
                float(st.G) if st.G is not None else None,
                float(st.S_G_N) if st.S_G_N is not None else None,
                float(st.BT) if st.BT is not None else None,
                st.tradable_days,
                float(st.BMJ) if st.BMJ is not None else None,
            ])

        # Charts
        if len(stats) > 2:
            # x axis dates in col A, series in columns E, F, H
            data_rows = len(stats) + 1  # header row + data
            cats = Reference(sh, min_col=1, min_row=2, max_row=data_rows)

            def add_chart(col_idx, title, anchor):
                chart = LineChart()
                chart.title = title
                chart.y_axis.title = "%"
                chart.x_axis.title = "Date"
                data = Reference(sh, min_col=col_idx, min_row=1, max_row=data_rows)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 8
                chart.width = 20
                sh.add_chart(chart, anchor)

            add_chart(5, "S_G_N (%)", "J2")
            add_chart(6, "BT (%)", "J18")
            add_chart(8, "BMJ (%)", "J34")

        for col in range(1, 9):
            sh.column_dimensions[get_column_letter(col)].width = 16

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
