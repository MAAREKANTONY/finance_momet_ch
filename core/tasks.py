from celery import shared_task
from decimal import Decimal
from datetime import datetime, date
from django.conf import settings
from django.core.mail import EmailMultiAlternatives

import hashlib
from datetime import timedelta
from django.db.models import Max, Q

from .models import Symbol, Scenario, DailyBar, DailyMetric, Alert, EmailRecipient, EmailSettings, AlertDefinition, GameScenario
from .models import Backtest
from .models import ProcessingJob
from .exports import build_scenario_workbook_write_only
from .services.provider_twelvedata import TwelveDataClient
from .services.calculations import compute_for_symbol_scenario
from .services.global_momentum import build_global_momentum_regime_by_date, GLOBAL_MOMENTUM_CODES
from .services.calculations_fast import compute_full_for_symbol_scenario

from django.utils import timezone
import json
import csv
import io
import zipfile
from pathlib import Path


class JobCancelled(Exception):
    """Raised when a ProcessingJob has cancel_requested=True."""


class JobKilled(Exception):
    """Raised when a ProcessingJob has kill_requested=True."""


def _job_checkpoint(job: ProcessingJob | None, *, heartbeat: bool = True) -> None:
    """Cooperative cancellation + heartbeat.

    - Refresh flags from DB
    - Update heartbeat_at periodically (cheap single-row update)
    - Raise JobCancelled/JobKilled when requested
    """
    if job is None:
        return
    # Refresh only the fields we need
    job.refresh_from_db(fields=["cancel_requested", "kill_requested", "status"])
    if job.kill_requested:
        raise JobKilled("kill requested")
    if job.cancel_requested:
        raise JobCancelled("cancel requested")
    if heartbeat:
        ProcessingJob.objects.filter(id=job.id).update(heartbeat_at=timezone.now())

def parse_date(s: str) -> date:
    return datetime.fromisoformat(s.replace("Z","")).date()

def desired_outputsize_years(years: int) -> int:
    # Roughly 252 trading days / year. Add buffer.
    if years <= 0:
        return 260
    return min(5000, years * 260)



def _parse_int_or_none(x):
    """Parse TwelveData numeric strings to int; return None if empty/invalid."""
    if x is None:
        return None
    if isinstance(x, int):
        return x
    s = str(x).strip()
    if s == "" or s.lower() == "null":
        return None
    try:
        return int(Decimal(s))
    except Exception:
        return None

def _fetch_daily_bars_for_symbols(*, symbol_qs, outputsize: int, force_full: bool = False, job: ProcessingJob | None = None) -> dict:
    """Fetch/update daily bars for a queryset of Symbol.

    Returns basic stats: {"symbols":..., "bars":...}.
    """
    client = TwelveDataClient()
    symbols = list(symbol_qs)
    bars_written = 0

    today = timezone.now().date()

    for sym in symbols:
        # Cooperative cancel/kill + heartbeat
        _job_checkpoint(job)
        exchange = sym.exchange or getattr(settings, "DEFAULT_EXCHANGE", "")
        try:
            # Delta fetch by default: if we already have bars, only request dates after the last stored bar.
            # This avoids re-downloading years of history each day.
            start_date = None
            if not force_full:
                last_date = DailyBar.objects.filter(symbol=sym).aggregate(Max("date")).get("date__max")
                if last_date:
                    start = last_date + timedelta(days=1)
                    if start <= today:
                        start_date = start.isoformat()
                    else:
                        # Already up to date.
                        continue

            values = client.time_series_daily(
                sym.ticker,
                exchange=exchange,
                outputsize=outputsize,
                start_date=start_date,
            )
        except Exception as e:
            print(f"[fetch] error {sym}: {e}")
            continue

        if not values:
            continue

        values_sorted = sorted(values, key=lambda v: v.get("datetime"))

        # Build new bars in memory and insert in bulk.
        new_bars = []
        for v in values_sorted:
            try:
                d = parse_date(v["datetime"])
                o = Decimal(v["open"]); h = Decimal(v["high"]); l = Decimal(v["low"]); c = Decimal(v["close"])
                vol = _parse_int_or_none(v.get("volume"))
            except Exception:
                continue

            if force_full:
                # Legacy behavior: upsert all rows (slower but keeps ability to refresh history).
                DailyBar.objects.update_or_create(
                    symbol=sym,
                    date=d,
                    defaults={"open": o, "high": h, "low": l, "close": c, "volume": vol, "source": "twelvedata"},
                )
                bars_written += 1
            else:
                # Delta mode: insert only new rows.
                new_bars.append(
                    DailyBar(
                        symbol=sym,
                        date=d,
                        open=o,
                        high=h,
                        low=l,
                        close=c,
                        volume=vol,
                        source="twelvedata",
                    )
                )

        if not force_full and new_bars:
            DailyBar.objects.bulk_create(new_bars, ignore_conflicts=True, batch_size=2000)
            bars_written += len(new_bars)

        # Update change_* for the latest bar (cheap and keeps UI consistent).
        last_two = list(DailyBar.objects.filter(symbol=sym).order_by("-date")[:2])
        if len(last_two) >= 2 and last_two[1].close:
            last_bar, prev_bar = last_two[0], last_two[1]
            change_amount = last_bar.close - prev_bar.close
            change_pct = (change_amount / prev_bar.close) * Decimal("100") if prev_bar.close != 0 else None
            DailyBar.objects.filter(id=last_bar.id).update(change_amount=change_amount, change_pct=change_pct)

    return {"symbols": len(symbols), "bars": bars_written, "force_full": bool(force_full)}


def _compute_metrics_for_scenario(*, symbols_qs, scenario: Scenario, recompute_all: bool = False, job: ProcessingJob | None = None) -> dict:
    """Compute DailyMetric + Alert for a given scenario and subset of symbols.

    Safety rule:
    - Scenario.history_years limits the *stored/computed* business window.
    - We still keep a technical lookback buffer before that window so indicators remain stable.
    - This avoids recomputing 10+ years when the scenario only asks for a shorter history.
    """
    symbols = list(symbols_qs)

    # Canonical signature of indicator parameters (stable across Scenario/GameScenario).
    cur_hash = indicator_signature(scenario)
    needs_full = recompute_all or (scenario.last_computed_config_hash and scenario.last_computed_config_hash != cur_hash)

    if needs_full:
        print(f"[compute] full recompute scenario={scenario.id} {scenario.name}")
        scenario.last_full_recompute_at = timezone.now()
        scenario.save(update_fields=["last_full_recompute_at"])

    n1 = int(scenario.n1 or 0)
    n2 = int(scenario.n2 or 0)
    n3 = int(scenario.n3 or 0)
    n4 = int(scenario.n4 or 0)
    n5 = int(getattr(scenario, 'n5', 0) or 0)
    k2j = int(getattr(scenario, 'k2j', 0) or 0)
    # K2f requires enough history to compute N5 variations + K2J smoothing.
    n5f3 = int(getattr(scenario, 'n5f3', 0) or 0)
    nampL3 = int(getattr(scenario, 'nampL3', 0) or 0)
    periodeL3 = int(getattr(scenario, 'periodeL3', 0) or 0)
    npente = int(getattr(scenario, 'npente', 0) or 0)
    npente_basse = int(getattr(scenario, 'npente_basse', 0) or 0)
    # Kf3 needs Mf1/Xf1: n5f3 + n5f3/2, plus amp window, plus slope window (can expand but we clamp to 5000).
    lookback_kf3 = max((n5f3 + max(1, n5f3 // 2) + 5) if n5f3 else 0, (nampL3 + 5) if nampL3 else 0, (periodeL3 + 5) if periodeL3 else 0)

    lookback_trading = max((n1 + n2 + 5), (n3 + n4 + 5), (n1 + 5), (n5 + k2j + 5), lookback_kf3, npente + 5, npente_basse + 5, 20)
    buffer_days = lookback_trading * 2 + 10
    history_years = max(1, int(getattr(scenario, 'history_years', 2) or 2))
    approx_business_window_days = int(history_years * 366)

    computed_rows = 0
    for sym in symbols:
        _job_checkpoint(job)
        try:
            sym_last_bar_date = DailyBar.objects.filter(symbol=sym).aggregate(m=Max("date"))["m"]
            if not sym_last_bar_date:
                continue

            business_start = sym_last_bar_date - timedelta(days=approx_business_window_days)
            technical_start = business_start - timedelta(days=buffer_days)

            # Always prune rows older than the technical window.
            # This keeps enough history for stable indicators while respecting history_years.
            Alert.objects.filter(symbol=sym, scenario=scenario, date__lt=technical_start).delete()
            DailyMetric.objects.filter(symbol=sym, scenario=scenario, date__lt=technical_start).delete()

            if needs_full:
                # Full recompute is scoped to the technical window only.
                Alert.objects.filter(scenario=scenario, symbol=sym).delete()
                DailyMetric.objects.filter(scenario=scenario, symbol=sym).delete()
                bars = (
                    DailyBar.objects.filter(symbol=sym, date__gte=technical_start)
                    .order_by("date")
                    .only("date", "open", "high", "low", "close")
                )
                m_written, a_written = compute_full_for_symbol_scenario(symbol=sym, scenario=scenario, bars=bars)
                computed_rows += m_written
                continue

            # Incremental recompute: only rebuild the recent technical tail, never the whole history.
            last_date = DailyMetric.objects.filter(symbol=sym, scenario=scenario).aggregate(m=Max("date"))["m"]
            if last_date:
                start = max(last_date - timedelta(days=buffer_days), technical_start)
                Alert.objects.filter(symbol=sym, scenario=scenario, date__gte=start).delete()
                DailyMetric.objects.filter(symbol=sym, scenario=scenario, date__gte=start).delete()
            else:
                start = technical_start

            bars_qs = DailyBar.objects.filter(symbol=sym, date__gte=start).order_by("date")

            _i = 0
            for d in bars_qs.values_list("date", flat=True):
                # Throttled checkpoints to avoid DB chatter on large histories.
                # Still responsive enough for manual cancels.
                _i += 1
                if job is not None and (_i % 200 == 0):
                    _job_checkpoint(job)
                compute_for_symbol_scenario(sym, scenario, d)
                computed_rows += 1
        except Exception as e:
            print(f"[compute] error {sym} {scenario}: {e}")
            continue

    try:
        _enrich_alerts_with_global_momentum(scenario=scenario, start_date=None)
    except Exception:
        pass

    Scenario.objects.filter(id=scenario.id).update(last_computed_config_hash=cur_hash)
    return {"symbols": len(symbols), "rows": computed_rows, "full": bool(needs_full)}




def _enrich_alerts_with_global_momentum(*, scenario: Scenario, start_date=None) -> int:
    """Append the current global momentum regime code to existing alert rows.

    Safe/additive behavior:
    - does not create standalone alert rows when no local alert exists
    - only enriches rows already produced by per-ticker alert calculations
    - removes stale GM_* codes before re-appending the current regime code
    """
    dm_qs = DailyMetric.objects.filter(scenario=scenario).order_by("symbol_id", "date")
    if start_date is not None:
        dm_qs = dm_qs.filter(date__gte=start_date)
    rows = list(dm_qs.values("symbol_id", "date", "P"))
    if not rows:
        return 0

    metrics_by_ticker: dict[int, dict] = {}
    for row in rows:
        metrics_by_ticker.setdefault(row["symbol_id"], {})[row["date"]] = row.get("P")

    nglobal = int(getattr(scenario, "nglobal", 20) or 20)
    regime_by_date = build_global_momentum_regime_by_date(metrics_by_ticker, nglobal=nglobal)
    if not regime_by_date:
        return 0

    alerts_qs = Alert.objects.filter(scenario=scenario)
    if start_date is not None:
        alerts_qs = alerts_qs.filter(date__gte=start_date)

    updated = 0
    batch = []
    for obj in alerts_qs.only("id", "date", "alerts"):
        regime = regime_by_date.get(obj.date)
        if not regime:
            continue
        codes = [c.strip() for c in str(obj.alerts or '').split(',') if c.strip()]
        codes = [c for c in codes if c not in GLOBAL_MOMENTUM_CODES]
        if regime not in codes:
            codes.append(regime)
        new_alerts = ','.join(codes)
        if new_alerts != (obj.alerts or ''):
            obj.alerts = new_alerts
            batch.append(obj)
            updated += 1
            if len(batch) >= 2000:
                Alert.objects.bulk_update(batch, ["alerts"], batch_size=2000)
                batch = []
    if batch:
        Alert.objects.bulk_update(batch, ["alerts"], batch_size=2000)
    return updated

def _indicator_params_from_scenario_like(obj) -> dict:
    """Canonical dict of parameters that affect indicator computations.

    IMPORTANT:
    - Must include every parameter that changes the math.
    - Must be stable across Scenario and GameScenario objects.
    - Must not include metadata (name, description, ids, timestamps).
    """
    # Normalize Decimals/None/ints as strings to make hash stable across DB/Python types.
    def norm(x):
        if x is None:
            return None
        try:
            # Decimal, int, str... -> string form
            return str(x)
        except Exception:
            return repr(x)

    fields = [
        # Common core
        "a", "b", "c", "d", "e", "vc", "fl",
        "n1", "n2", "n3", "n4",
        # K2f
        "n5", "k2j", "cr",
        # Kf3
        "n5f3", "crf3", "nampL3", "baseL3", "periodeL3",
        # SUM_SLOPE / SPa-SPv
        "npente", "slope_threshold", "npente_basse", "slope_threshold_basse",
        # V line
        "m_v",
    ]
    return {f: norm(getattr(obj, f, None)) for f in fields}


def indicator_signature(obj) -> str:
    """SHA-256 of a canonical JSON payload of indicator parameters."""
    payload = json.dumps(_indicator_params_from_scenario_like(obj), sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _buffer_days_for_scenario(scenario: Scenario) -> int:
    """Compute the same rolling window buffer used by incremental compute."""
    n1 = int(scenario.n1 or 0)
    n2 = int(scenario.n2 or 0)
    n3 = int(scenario.n3 or 0)
    n4 = int(scenario.n4 or 0)
    n5 = int(getattr(scenario, 'n5', 0) or 0)
    k2j = int(getattr(scenario, 'k2j', 0) or 0)
    n5f3 = int(getattr(scenario, 'n5f3', 0) or 0)
    nampL3 = int(getattr(scenario, 'nampL3', 0) or 0)
    periodeL3 = int(getattr(scenario, 'periodeL3', 0) or 0)
    npente = int(getattr(scenario, 'npente', 0) or 0)
    npente_basse = int(getattr(scenario, 'npente_basse', 0) or 0)

    lookback_kf3 = max(
        (n5f3 + max(1, n5f3 // 2) + 5) if n5f3 else 0,
        (nampL3 + 5) if nampL3 else 0,
        (periodeL3 + 5) if periodeL3 else 0,
    )
    lookback_sum_slope = max((npente + 5) if npente else 0, (npente_basse + 5) if npente_basse else 0)
    lookback_trading = max((n1 + n2 + 5), (n3 + n4 + 5), (n1 + 5), (n5 + k2j + 5), lookback_kf3, lookback_sum_slope, 20)
    return int(lookback_trading * 2 + 10)


def _clone_metrics_and_alerts(
    *,
    from_scenario: Scenario,
    to_scenario: Scenario,
    symbols: list[Symbol],
    start_date,
    job: ProcessingJob | None = None,
) -> None:
    """Clone metrics+alerts from one scenario to another for a given date window.

    This enables safe deduplication when the indicator parameters are identical.
    """
    metric_fields = [
        "date",
        "P", "M", "M1", "X", "X1", "T", "Q", "S",
        "K1", "K1f", "K2f", "K2f_pre", "Kf2bis",
        "Kf3", "K2", "K3", "K4",
        "V_pre", "V_line",
        "V", "slope_P", "sum_slope", "slope_vrai", "sum_slope_basse", "slope_vrai_basse", "sum_pos_P", "nb_pos_P", "ratio_P", "amp_h",
    ]

    for sym in symbols:
        _job_checkpoint(job, heartbeat=False)

        # Delete the target window first (keep behavior aligned with incremental compute).
        if start_date is not None:
            DailyMetric.objects.filter(symbol=sym, scenario=to_scenario, date__gte=start_date).delete()
            Alert.objects.filter(symbol=sym, scenario=to_scenario, date__gte=start_date).delete()
            src_metrics_qs = DailyMetric.objects.filter(symbol=sym, scenario=from_scenario, date__gte=start_date)
            src_alerts_qs = Alert.objects.filter(symbol=sym, scenario=from_scenario, date__gte=start_date)
        else:
            DailyMetric.objects.filter(symbol=sym, scenario=to_scenario).delete()
            Alert.objects.filter(symbol=sym, scenario=to_scenario).delete()
            src_metrics_qs = DailyMetric.objects.filter(symbol=sym, scenario=from_scenario)
            src_alerts_qs = Alert.objects.filter(symbol=sym, scenario=from_scenario)

        # Copy metrics
        m_rows = list(src_metrics_qs.values(*metric_fields))
        if m_rows:
            objs = [DailyMetric(symbol=sym, scenario=to_scenario, **row) for row in m_rows]
            DailyMetric.objects.bulk_create(objs, batch_size=5000)

        # Copy alerts
        a_rows = list(src_alerts_qs.values("date", "alerts"))
        if a_rows:
            aobjs = [Alert(symbol=sym, scenario=to_scenario, date=row["date"], alerts=row["alerts"]) for row in a_rows]
            Alert.objects.bulk_create(aobjs, batch_size=5000)

    # Keep config hash aligned to avoid unexpected full recompute on next runs.
    Scenario.objects.filter(id=to_scenario.id).update(last_computed_config_hash=from_scenario.last_computed_config_hash)


@shared_task
def fetch_daily_bars_task():
    symbols = Symbol.objects.filter(active=True).all()
    max_years = Scenario.objects.filter(active=True).order_by("-history_years").values_list("history_years", flat=True).first() or 2
    outputsize = desired_outputsize_years(int(max_years))

    stats = _fetch_daily_bars_for_symbols(symbol_qs=symbols, outputsize=outputsize)
    return f"ok outputsize={outputsize} symbols={stats.get('symbols')} bars={stats.get('bars')}"

@shared_task
def compute_metrics_task(recompute_all: bool = False):
    """Compute metrics and alerts.

    Default behavior is **incremental** (recompute the recent window + new days).
    If scenario variables changed since last compute, we do a **full recompute** for that scenario.
    If recompute_all=True, force full recompute for all scenarios.
    """
    symbols = Symbol.objects.filter(active=True).all()
    scenarios = Scenario.objects.filter(active=True).all()

    for scenario in scenarios:
        _compute_metrics_for_scenario(symbols_qs=symbols, scenario=scenario, recompute_all=recompute_all)

    return "ok"


@shared_task
def _noop():
    return "ok"


@shared_task(bind=True)
def fetch_daily_bars_job_task(self, *, symbol_ids=None, scenario_id=None, force_full: bool = False, backtest_id=None, user_id=None, job_id=None):
    """Tracked job wrapper around DailyBar fetching.

    If symbol_ids is None -> fetch for all active symbols.
    If scenario_id provided -> outputsize based on scenario.history_years.
    force_full=True keeps legacy upsert behavior to refresh historical rows.
    """
    job = None
    if job_id:
        job = ProcessingJob.objects.filter(id=job_id).first()
        if job:
            job.status = ProcessingJob.Status.RUNNING
            job.task_id = getattr(self.request, "id", "") or ""
            job.started_at = timezone.now()
            job.save(update_fields=["status", "task_id", "started_at"])
    if job is None:
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.FETCH_BARS,
            status=ProcessingJob.Status.RUNNING,
            task_id=getattr(self.request, "id", "") or "",
            backtest_id=backtest_id,
            scenario_id=scenario_id,
            created_by_id=user_id,
            started_at=timezone.now(),
        )
    try:
        # Initial checkpoint + heartbeat
        _job_checkpoint(job)
        if symbol_ids:
            symbol_qs = Symbol.objects.filter(id__in=list(symbol_ids))
        else:
            symbol_qs = Symbol.objects.filter(active=True)

        years = None
        if scenario_id:
            years = Scenario.objects.filter(id=scenario_id).values_list("history_years", flat=True).first()
        if years is None:
            years = Scenario.objects.filter(active=True).order_by("-history_years").values_list("history_years", flat=True).first() or 2
        outputsize = desired_outputsize_years(int(years))

        stats = _fetch_daily_bars_for_symbols(symbol_qs=symbol_qs, outputsize=outputsize, force_full=bool(force_full), job=job)
        job.status = ProcessingJob.Status.DONE
        job.message = (
            f"Fetched bars: symbols={stats.get('symbols')} bars={stats.get('bars')} "
            f"outputsize={outputsize} force_full={bool(force_full)}"
        )
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return job.message
    except JobCancelled:
        job.status = ProcessingJob.Status.CANCELLED
        job.message = (job.message or "") + "\nCancelled by user."
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return "cancelled"
    except JobKilled:
        job.status = ProcessingJob.Status.KILLED
        job.message = (job.message or "") + "\nKilled by user."
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return "killed"
    except Exception as e:
        job.status = ProcessingJob.Status.FAILED
        job.error = str(e)
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "error", "finished_at"])
        raise


@shared_task(bind=True)
def compute_metrics_job_task(self, *, scenario_id, symbol_ids=None, recompute_all=False, backtest_id=None, user_id=None, job_id=None):
    """Tracked job wrapper around metrics computation."""
    job = None
    if job_id:
        job = ProcessingJob.objects.filter(id=job_id).first()
        if job:
            job.status = ProcessingJob.Status.RUNNING
            job.task_id = getattr(self.request, "id", "") or ""
            job.started_at = timezone.now()
            job.save(update_fields=["status", "task_id", "started_at"])
    if job is None:
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.COMPUTE_METRICS,
            status=ProcessingJob.Status.RUNNING,
            task_id=getattr(self.request, "id", "") or "",
            backtest_id=backtest_id,
            scenario_id=scenario_id,
            created_by_id=user_id,
            started_at=timezone.now(),
        )
    try:
        _job_checkpoint(job)
        scenario = Scenario.objects.get(id=scenario_id)
        # Scoping rules (no regression for legacy flows):
        # - If explicit symbol_ids are provided (e.g., from a Backtest universe snapshot), compute only those.
        # - Otherwise, when computing a single scenario from UI, compute only the symbols attached to that scenario.
        #   Fallback to all active symbols only if the scenario has no explicit universe.
        if symbol_ids:
            symbols_qs = Symbol.objects.filter(id__in=list(symbol_ids))
            scope_note = f"explicit_ids={len(list(symbol_ids))}"
        else:
            # When no explicit symbol_ids are provided, we interpret this job as "compute this scenario".
            # To avoid surprise recomputes across the whole universe, we scope strictly to the scenario symbols.
            scenario_symbols_qs = scenario.symbols.filter(active=True)
            if not scenario_symbols_qs.exists():
                job.status = ProcessingJob.Status.DONE
                job.message = "No symbols linked to this scenario (nothing to compute)."
                job.finished_at = timezone.now()
                job.save(update_fields=["status", "message", "finished_at"])
                return job.message
            symbols_qs = scenario_symbols_qs
            scope_note = "scenario_symbols"

        stats = _compute_metrics_for_scenario(
            symbols_qs=symbols_qs,
            scenario=scenario,
            recompute_all=bool(recompute_all),
            job=job,
        )
        job.status = ProcessingJob.Status.DONE
        job.message = (
            f"Computed metrics: scope={scope_note} scenario={scenario.id} "
            f"symbols={stats.get('symbols')} rows={stats.get('rows')} full={stats.get('full')}"
        )
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return job.message
    except JobCancelled:
        job.status = ProcessingJob.Status.CANCELLED
        job.message = (job.message or "") + "\nCancelled by user."
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return "cancelled"
    except JobKilled:
        job.status = ProcessingJob.Status.KILLED
        job.message = (job.message or "") + "\nKilled by user."
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return "killed"
    except Exception as e:
        job.status = ProcessingJob.Status.FAILED
        job.error = str(e)
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "error", "finished_at"])
        raise


@shared_task(bind=True)
def compute_metrics_all_job_task(self, recompute_all: bool = False, user_id=None):
    """Tracked job wrapper for a full compute across all active scenarios."""
    job = ProcessingJob.objects.create(
        job_type=ProcessingJob.JobType.COMPUTE_METRICS,
        status=ProcessingJob.Status.RUNNING,
        task_id=getattr(self.request, "id", "") or "",
        created_by_id=user_id,
        started_at=timezone.now(),
        message=f"compute_all recompute_all={bool(recompute_all)}",
    )
    try:
        msg = compute_metrics_task(bool(recompute_all))
        job.status = ProcessingJob.Status.DONE
        job.message = f"{job.message} -> {msg}"
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return msg
    except Exception as e:
        job.status = ProcessingJob.Status.FAILED
        job.error = str(e)
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "error", "finished_at"])
        raise

@shared_task
def send_daily_alerts_task():
    """
    Sends a single recap email for the most recent alert date.
    Includes RATIO_P and AMP_H trend indicators per (symbol, scenario).
    """
    last = Alert.objects.order_by("-date").first()
    if not last:
        return "no-alerts"

    alert_date = last.date
    alerts = Alert.objects.filter(date=alert_date).select_related("symbol", "scenario").order_by("scenario__name", "symbol__ticker")
    if not alerts.exists():
        return "no-alerts-today"

    recipients = list(EmailRecipient.objects.filter(active=True).values_list("email", flat=True))
    if not recipients:
        return "no-recipients"

    def fmt_pct(x):
        """Formats values that are already stored as percentages (0-100)."""
        if x is None:
            return "—"
        try:
            return f"{Decimal(x):.2f}%"
        except Exception:
            return "—"

    def fmt_num(x):
        if x is None:
            return "—"
        try:
            return f"{Decimal(x):.6f}"
        except Exception:
            return "—"

    rows = []
    for a in alerts:
        m = DailyMetric.objects.filter(symbol=a.symbol, scenario=a.scenario, date=alert_date).first()
        ratio_p = fmt_pct(getattr(m, "ratio_P", None) if m else None)
        amp_h = fmt_pct(getattr(m, "amp_h", None) if m else None)
        rows.append(
            f"<tr>"
            f"<td>{a.date}</td>"
            f"<td>{a.symbol.ticker}</td>"
            f"<td>{a.scenario.name}</td>"
            f"<td>{a.alerts}</td>"
            f"<td>{ratio_p}</td>"
            f"<td>{amp_h}</td>"
            f"</tr>"
        )

    html = f"""
    <h3>Alertes bourse - {alert_date}</h3>
    <p>Un seul email récapitulatif pour tous les scénarios.</p>
    <table border="1" cellpadding="6" cellspacing="0">
      <thead>
        <tr>
          <th>Date</th><th>Action</th><th>Scénario</th><th>Alertes</th><th>RATIO_P</th><th>AMP_H</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows)}
      </tbody>
    </table>
    """

    msg = EmailMultiAlternatives(
        subject=f"Stock Alerts - {alert_date}",
        body=f"Alertes bourse - {alert_date}",
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        to=recipients,
    )
    msg.attach_alternative(html, "text/html")
    msg.send()
    return "sent"


def _send_alert_definition_email(defn: AlertDefinition, alert_date):
    """Send one email for a specific AlertDefinition.

    NO-REGRESSION: this reads from the existing `Alert` table; it does not
    change how alerts are computed.
    """
    recipients = list(defn.recipients.filter(active=True).values_list("email", flat=True))
    if not recipients:
        return "no-recipients"

    qs = Alert.objects.filter(date=alert_date).select_related("symbol", "scenario")
    # Scenario filter (empty => all)
    scenario_ids = list(defn.scenarios.values_list("id", flat=True))
    if scenario_ids:
        qs = qs.filter(scenario_id__in=scenario_ids)

    codes = defn.get_codes_list()
    if codes:
        q = Q()
        for c in codes:
            q |= Q(alerts__icontains=c)
        qs = qs.filter(q)

    alerts = list(qs.order_by("scenario__name", "symbol__ticker"))
    if not alerts:
        return "no-alerts"

    def fmt_pct(x):
        if x is None:
            return "—"
        try:
            return f"{Decimal(x):.2f}%"
        except Exception:
            return "—"

    rows = []
    for a in alerts:
        m = DailyMetric.objects.filter(symbol=a.symbol, scenario=a.scenario, date=alert_date).first()
        ratio_p = fmt_pct(getattr(m, "ratio_P", None) if m else None)
        amp_h = fmt_pct(getattr(m, "amp_h", None) if m else None)
        rows.append(
            f"<tr>"
            f"<td>{a.date}</td>"
            f"<td>{a.symbol.ticker}</td>"
            f"<td>{a.scenario.name}</td>"
            f"<td>{a.alerts}</td>"
            f"<td>{ratio_p}</td>"
            f"<td>{amp_h}</td>"
            f"</tr>"
        )

    html = f"""
    <h3>Stock Alerts - {alert_date}</h3>
    <p><strong>{defn.name}</strong></p>
    <p class=\"muted\">{defn.description or ''}</p>
    <table border=\"1\" cellpadding=\"6\" cellspacing=\"0\">
      <thead>
        <tr>
          <th>Date</th><th>Action</th><th>Scénario</th><th>Alertes</th><th>RATIO_P</th><th>AMP_H</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows)}
      </tbody>
    </table>
    """

    msg = EmailMultiAlternatives(
        subject=f"Stock Alerts - {defn.name} - {alert_date}",
        body=f"Stock Alerts - {defn.name} - {alert_date}",
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        to=recipients,
    )
    msg.attach_alternative(html, "text/html")
    msg.send()
    return "sent"

@shared_task
def send_alert_definition_now_task(definition_id: int):
    """Send one configured alert definition immediately.

    This is used by the UI action 'Envoyer'.
    NO-REGRESSION: reads from existing Alert rows and does not change computation.
    """
    last = Alert.objects.order_by("-date").first()
    if not last:
        return "no-alert-data"
    defn = AlertDefinition.objects.filter(id=definition_id).prefetch_related("scenarios", "recipients").first()
    if not defn:
        return "not-found"
    return _send_alert_definition_email(defn, last.date)




@shared_task(bind=True)
def send_alert_definition_job_task(self, *, alert_definition_id: int, user_id=None, job_id=None):
    """Tracked job wrapper to send a single AlertDefinition immediately."""
    job = None
    if job_id:
        job = ProcessingJob.objects.filter(id=job_id).first()
        if job:
            job.status = ProcessingJob.Status.RUNNING
            job.task_id = getattr(self.request, "id", "") or ""
            job.started_at = timezone.now()
            job.message = f"En cours (envoi alerte {alert_definition_id})"
            job.save(update_fields=["status", "task_id", "started_at", "message"])

    try:
        # Reuse the existing sending logic synchronously so we can update the job status reliably.
        send_alert_definition_now_task(alert_definition_id)

        if job:
            job.status = ProcessingJob.Status.DONE
            job.finished_at = timezone.now()
            job.message = f"Terminé (envoi alerte {alert_definition_id})"
            job.save(update_fields=["status", "finished_at", "message"])
        return {"status": "ok"}

    except Exception as e:
        if job:
            job.status = ProcessingJob.Status.FAILED
            job.finished_at = timezone.now()
            job.message = f"Erreur (envoi alerte {alert_definition_id}): {e}"
            job.save(update_fields=["status", "finished_at", "message"])
        raise
@shared_task
def check_and_send_scheduled_alerts_task():
    """Runs every hour (minute=0), with Redis lock to prevent overlaps.

    Legacy behavior (NO-REGRESSION):
    - Uses `EmailSettings` to send ONE recap email for all scenarios.

    Additive behavior:
    - Also evaluates user-defined `AlertDefinition` rules (scenarios + lines + recipients + schedule).
    """
    from django.utils import timezone as dj_tz
    # --- Anti-overlap lock (prevents queue backlog if beat fires while a previous run is still running) ---
    acquired_lock = False
    try:
        from django.conf import settings as dj_settings
        import redis
        r = redis.Redis.from_url(dj_settings.CELERY_BROKER_URL)
        # Keep the lock slightly below the schedule interval (1h) to tolerate crashes without permanent lock.
        acquired_lock = bool(r.set('lock:check_scheduled_alerts', '1', nx=True, ex=55 * 60))
    except Exception:
        # If Redis lock fails, we proceed (no regression).
        acquired_lock = True
    if not acquired_lock:
        return 'locked_skip'

    email_cfg = EmailSettings.get_solo()

    try:
        import zoneinfo
        tz = zoneinfo.ZoneInfo(email_cfg.timezone or "Asia/Jerusalem")
    except Exception:
        tz = dj_tz.get_current_timezone()

    now = dj_tz.now().astimezone(tz)
    today = now.date()

    results = []

    # --- Legacy global email (unchanged) ---
    if email_cfg.last_sent_date != today:
        if now.hour == int(email_cfg.send_hour) and now.minute == int(email_cfg.send_minute):
            send_daily_alerts_task.delay()
            email_cfg.last_sent_date = today
            email_cfg.save(update_fields=["last_sent_date", "updated_at"])
            results.append("global_sent")
        else:
            results.append("global_not_due")
    else:
        results.append("global_already_sent")

    # --- Additive: user-defined alert definitions ---
    last = Alert.objects.order_by("-date").first()
    if not last:
        results.append("no_alert_data")
        return ";".join(results)

    alert_date = last.date
    defs = AlertDefinition.objects.filter(is_active=True).prefetch_related("scenarios", "recipients")
    for d in defs:
        try:
            import zoneinfo
            dtz = zoneinfo.ZoneInfo(d.timezone or "Asia/Jerusalem")
        except Exception:
            dtz = tz

        dnow = dj_tz.now().astimezone(dtz)
        dtoday = dnow.date()

        if d.last_sent_date == dtoday:
            continue

        if dnow.hour == int(d.send_hour) and dnow.minute == int(d.send_minute):
            _send_alert_definition_email(d, alert_date)
            d.last_sent_date = dtoday
            d.save(update_fields=["last_sent_date", "updated_at"])
            results.append(f"def_sent#{d.id}")

    return ";".join(results)

def run_game_scenario_task(game_id: int, force_fetch: bool = False, force_recompute: bool = False):
    """Run one GameScenario end-to-end."""
    from core.services.game_scenarios.runner import run_game_scenario_now

    return run_game_scenario_now(game_id, force_fetch=force_fetch, force_recompute=force_recompute)


def _ensure_game_engine_scenario(game: GameScenario) -> Scenario:
    """Create/update the internal Scenario used by a GameScenario.

    Duplicates the logic from game runner to avoid circular imports.
    """
    sc = game.engine_scenario
    if sc is None:
        sc = Scenario(
            name=f"[GAME] {game.name}",
            description=f"Auto-generated scenario for GameScenario #{game.id}",
            active=False,
            is_default=False,
        )

    for f in [
        "a", "b", "c", "d", "e", "vc", "fl",
        "n1", "n2", "n3", "n4",
        "n5", "k2j", "cr",
        "n5f3", "crf3", "nampL3", "baseL3", "periodeL3",
        "npente", "slope_threshold", "npente_basse", "slope_threshold_basse",
        "m_v",
    ]:
        setattr(sc, f, getattr(game, f))

    sc.name = f"[GAME] {game.name}"
    sc.description = f"Auto-generated scenario for GameScenario #{game.id}"
    sc.active = False
    sc.is_default = False
    sc.save()

    if game.engine_scenario_id != sc.id:
        game.engine_scenario = sc
        game.save(update_fields=["engine_scenario", "updated_at"])
    return sc


@shared_task(bind=True)
def daily_system_refresh_job_task(self, *, user_id=None, job_id=None):
    """Daily end-to-end refresh (scheduled at 08:00).

    What it does (as requested):
      1) Fetch latest DailyBars for ALL active tickers (delta fetch, no full refresh)
      2) Compute indicators for ALL active Scenarios (incremental; no full recompute unless operator forces it manually)
      3) Compute Game tables by running ALL active GameScenarios (uses their incremental compute rules)

    Safety:
      - Redis lock prevents overlaps
      - ProcessingJob tracking + cooperative cancel/kill
    """
    job = None
    if job_id:
        job = ProcessingJob.objects.filter(id=job_id).first()
    if job is None:
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.COMPUTE_METRICS,
            status=ProcessingJob.Status.PENDING,
            task_id=getattr(self.request, "id", "") or "",
            created_by_id=user_id,
            message="Daily system refresh (scheduled)",
        )

    # --- Anti-overlap lock (daily) ---
    acquired_lock = False
    try:
        import redis
        r = redis.Redis.from_url(settings.CELERY_BROKER_URL)
        # 6 hours TTL is a safe upper bound for heavy runs.
        acquired_lock = bool(r.set('lock:daily_system_refresh', '1', nx=True, ex=6 * 60 * 60))
    except Exception:
        acquired_lock = True
    if not acquired_lock:
        job.status = ProcessingJob.Status.DONE
        job.started_at = timezone.now()
        job.finished_at = timezone.now()
        job.message = "Daily refresh skipped (lock already held)"
        job.save(update_fields=["status", "started_at", "finished_at", "message"])
        return "locked_skip"

    job.status = ProcessingJob.Status.RUNNING
    job.started_at = timezone.now()
    job.save(update_fields=["status", "started_at"])

    try:
        # 1) Fetch bars for ALL active symbols
        symbols = Symbol.objects.filter(active=True).order_by("ticker")

        # Outputsize heuristic: max(scenarios history_years, games study_days)
        max_years = Scenario.objects.filter(active=True).order_by("-history_years").values_list("history_years", flat=True).first() or 2
        max_study_days = GameScenario.objects.filter(active=True).order_by("-study_days").values_list("study_days", flat=True).first() or 0
        outputsize = max(desired_outputsize_years(int(max_years)), min(5000, int(max_study_days) + 400) if max_study_days else 0)
        outputsize = max(260, min(5000, int(outputsize)))

        job.message = f"Step1/3 Fetch DailyBars (delta) outputsize={outputsize}"
        job.save(update_fields=["message"])
        _fetch_daily_bars_for_symbols(symbol_qs=symbols, outputsize=outputsize, force_full=False, job=job)

        # 2) Compute metrics for ALL active scenarios + ALL active games' engine scenarios.
        #    Optimization: canonical signature (hash of indicator parameters) => compute once per signature,
        #    then clone metrics+alerts to other scenarios that share the exact same parameters.

        symbols_all = list(Symbol.objects.filter(active=True).order_by("ticker"))

        targets = []  # list of dicts: {"scenario": Scenario, "symbols": list[Symbol], "kind": str, "owner_id": int}

        # Regular scenarios
        for sc in Scenario.objects.filter(active=True).order_by("id"):
            sc_syms = list(sc.symbols.filter(active=True).order_by("ticker"))
            if not sc_syms:
                continue
            targets.append({"scenario": sc, "symbols": sc_syms, "kind": "scenario", "owner_id": sc.id})

        # Game engine scenarios (parameters can differ from regular scenarios)
        games = list(GameScenario.objects.filter(active=True).order_by("id"))
        for g in games:
            sc = _ensure_game_engine_scenario(g)
            targets.append({"scenario": sc, "symbols": symbols_all, "kind": "game", "owner_id": g.id})

        # Group by signature
        groups = {}
        for t in targets:
            sig = indicator_signature(t["scenario"])
            groups.setdefault(sig, []).append(t)

        done_groups = 0
        for sig, group_targets in groups.items():
            _job_checkpoint(job)
            done_groups += 1

            # Pick primary scenario for real computation
            primary = group_targets[0]["scenario"]

            # Union of symbols across targets for the primary compute
            union_ids = set()
            for t in group_targets:
                for s in t["symbols"]:
                    union_ids.add(s.id)
            union_symbols = list(Symbol.objects.filter(id__in=list(union_ids)).order_by("ticker"))

            job.message = (
                f"Step2/3 Compute metrics group={done_groups}/{len(groups)} "
                f"sig={sig[:10]} primary_scenario={primary.id} symbols={len(union_symbols)} targets={len(group_targets)}"
            )
            job.save(update_fields=["message"])

            # Compute once (incremental)
            _compute_metrics_for_scenario(symbols_qs=union_symbols, scenario=primary, recompute_all=False, job=job)

            # Ensure primary keeps its signature hash set
            Scenario.objects.filter(id=primary.id).update(last_computed_config_hash=sig)

            # Clone to the other scenarios in this group
            for t in group_targets[1:]:
                _job_checkpoint(job, heartbeat=False)
                to_sc = t["scenario"]
                to_syms = t["symbols"]
                buffer_days = _buffer_days_for_scenario(primary)

                # Determine start_date for cloning window based on target's last computed date.
                # If the target has no data, we clone full history for its symbol scope.
                last_dt = DailyMetric.objects.filter(scenario=to_sc, symbol__in=to_syms).aggregate(m=Max("date"))["m"]
                start_dt = (last_dt - timedelta(days=buffer_days)) if last_dt else None

                job.message = (
                    f"Step2/3 Clone metrics sig={sig[:10]} from={primary.id} to={to_sc.id} "
                    f"symbols={len(to_syms)} start={start_dt or 'ALL'}"
                )
                job.save(update_fields=["message"])

                _clone_metrics_and_alerts(
                    from_scenario=primary,
                    to_scenario=to_sc,
                    symbols=to_syms,
                    start_date=start_dt,
                    job=job,
                )

        done_sc = sum(1 for t in targets if t["kind"] == "scenario")

        # 3) Run ALL active games (compute today's tables).
        # We pass skip_metrics=True because metrics were refreshed in step2.
        from core.services.game_scenarios.runner import run_game_scenario_now, GameJobCancelled, GameJobKilled
        total_g = len(games)
        done_g = 0
        for gs in games:
            _job_checkpoint(job)
            job.message = f"Step3/3 Run game={gs.id} ({done_g+1}/{total_g})"
            job.save(update_fields=["message"])
            try:
                run_game_scenario_now(gs.id, force_fetch=False, force_recompute=False, skip_metrics=True, job=job)
            except (GameJobCancelled, GameJobKilled):
                raise
            done_g += 1

        job.status = ProcessingJob.Status.DONE
        job.finished_at = timezone.now()
        job.message = f"Daily refresh OK: symbols={symbols.count()} scenarios={done_sc} games={done_g}"
        job.save(update_fields=["status", "finished_at", "message"])
        return job.message

    except JobCancelled:
        job.status = ProcessingJob.Status.CANCELLED
        job.finished_at = timezone.now()
        job.message = (job.message or "") + "\nCancelled by user."
        job.save(update_fields=["status", "finished_at", "message"])
        return "cancelled"
    except JobKilled:
        job.status = ProcessingJob.Status.KILLED
        job.finished_at = timezone.now()
        job.message = (job.message or "") + "\nKilled by user."
        job.save(update_fields=["status", "finished_at", "message"])
        return "killed"
    except Exception as e:
        job.status = ProcessingJob.Status.FAILED
        job.finished_at = timezone.now()
        job.error = str(e)
        job.message = (job.message or "") + f"\nFAILED: {e}"
        job.save(update_fields=["status", "finished_at", "error", "message"])
        raise


def _maybe_run_scheduled_game_scenarios(now_dt):
    """Minute-level scheduler for GameScenarios.

    Uses settings:
      GAME_SCENARIO_RUN_HOUR (default 3)
      GAME_SCENARIO_RUN_MINUTE (default 5)

    Runs each active GameScenario at most once per day (based on today_results.date).
    """
    try:
        hour = int(getattr(settings, "GAME_SCENARIO_RUN_HOUR", 3))
        minute = int(getattr(settings, "GAME_SCENARIO_RUN_MINUTE", 5))
    except Exception:
        hour, minute = 3, 5

    if now_dt.hour != hour or now_dt.minute != minute:
        return "games_not_due"

    today = now_dt.date()
    qs = GameScenario.objects.filter(active=True)
    scheduled = 0
    for gs in qs:
        last = gs.today_results.get("date") if isinstance(gs.today_results, dict) else None
        if last == str(today):
            continue
        if (gs.last_run_status or "").lower() == "running":
            continue
        run_game_scenario_task.delay(gs.id)
        scheduled += 1
    # --- Additive: GameScenarios scheduler (daily BMD table) ---
    try:
        results.append(_maybe_run_scheduled_game_scenarios(now))
    except Exception as e:
        results.append(f"games_error:{e}")

    return f"games_scheduled:{scheduled}"



@shared_task(bind=True)
def run_game_scenario_job_task(self, *, game_id: int, force_fetch: bool = False, force_recompute: bool = False, user_id=None, job_id=None):
    """Tracked wrapper for GameScenario runs."""
    job = None
    if job_id:
        job = ProcessingJob.objects.filter(id=job_id).first()
        if job:
            job.status = ProcessingJob.Status.RUNNING
            job.task_id = getattr(self.request, "id", "") or ""
            job.started_at = timezone.now()
            job.message = f"En cours (run game {game_id})"
            job.save(update_fields=["status", "task_id", "started_at", "message"])
    try:
        from core.services.game_scenarios.runner import run_game_scenario_now
        run_game_scenario_now(game_id, force_fetch=force_fetch, force_recompute=force_recompute)

        if job:
            job.status = ProcessingJob.Status.DONE
            job.finished_at = timezone.now()
            job.message = f"Terminé (run game {game_id})"
            job.save(update_fields=["status", "finished_at", "message"])
        return {"status": "ok"}
    except Exception as e:
        if job:
            job.status = ProcessingJob.Status.FAILED
            job.finished_at = timezone.now()
            job.message = f"Erreur (run game {game_id}): {e}"
            job.save(update_fields=["status", "finished_at", "message"])
        raise
@shared_task
def run_backtest_task(backtest_id: int):
    # Lazy imports to avoid circular imports
    from .services.backtesting.prep import prepare_backtest_data
    from .services.backtesting.engine import run_backtest as engine_run_backtest
    """Run a backtest end-to-end (Feature 3).

    Steps:
    1) Mark Backtest as RUNNING
    2) Ensure prerequisite data exists (fetch bars / compute metrics) if missing
    3) Run the backtest engine (minimal implementation)
    4) Persist results JSON and mark DONE (or FAILED)
    """
    bt = Backtest.objects.filter(id=backtest_id).first()
    if not bt:
        return f"backtest {backtest_id} not found"

    Backtest.objects.filter(id=bt.id).update(status=Backtest.Status.RUNNING, error_message="")
    try:
        prep_report = prepare_backtest_data(bt)
        engine_result = engine_run_backtest(bt)
        results = engine_result.results

        # --- Optional (NO-REGRESSION) Parquet storage ---
        # Writes daily series to Parquet *in addition* to existing JSON results.
        # Enabled only when ENABLE_PARQUET_STORAGE=1.
        try:
            from .services.backtesting.parquet_storage import write_backtest_parquet_files

            write_backtest_parquet_files(bt, results)
        except Exception:
            # Never fail a backtest because of optional storage.
            pass

        results["prep"] = {
            "did_fetch_bars": prep_report.did_fetch_bars,
            "did_compute_metrics": prep_report.did_compute_metrics,
            "notes": prep_report.notes,
        }

        # --- Automatic results offload (NO-REGRESSION) ---
        # PostgreSQL JSONB has a hard limit (~256MB) for object elements.
        # Large universes (e.g., S&P 500) can exceed this when storing the full
        # per-ticker daily series inside Backtest.results. When the payload is
        # too large, we offload only the heavy `daily` arrays to files on disk
        # and keep lightweight pointers in JSON.
        try:
            from .services.backtesting.results_offload import offload_daily_series_if_needed

            results = offload_daily_series_if_needed(bt, results)
        except Exception:
            # Never fail a backtest because of optional offload logic.
            pass
        # Persist results JSON + portfolio tables (Feature 8)
        from django.db import transaction
        from .models import BacktestPortfolioDaily, BacktestPortfolioKPI

        portfolio = results.get("portfolio") or {}
        port_daily = portfolio.get("daily") or []
        port_kpi = portfolio.get("kpi") or {}

        with transaction.atomic():
            Backtest.objects.filter(id=bt.id).update(status=Backtest.Status.DONE, results=results, error_message="")

            # Replace portfolio daily rows
            BacktestPortfolioDaily.objects.filter(backtest_id=bt.id).delete()
            daily_objs = []
            for r in port_daily:
                try:
                    daily_objs.append(
                        BacktestPortfolioDaily(
                            backtest_id=bt.id,
                            date=r.get("date"),
                            global_cash=r.get("global_cash") or 0,
                            cash_allocated=r.get("cash_allocated") or 0,
                            positions_value=r.get("positions_value") or 0,
                            equity=r.get("equity") or 0,
                            invested=r.get("invested") or 0,
                            drawdown=r.get("drawdown") or 0,
                        )
                    )
                except Exception:
                    continue
            if daily_objs:
                BacktestPortfolioDaily.objects.bulk_create(daily_objs, batch_size=1000)

            # Upsert KPI row
            BacktestPortfolioKPI.objects.update_or_create(
                backtest_id=bt.id,
                defaults={
                    "capital_total": port_kpi.get("capital_total") or 0,
                    "invested_end": port_kpi.get("invested_end") or 0,
                    "equity_end": port_kpi.get("equity_end") or 0,
                    "bt_return": port_kpi.get("BT"),
                    "bmj_return": port_kpi.get("BMJ"),
                    "nb_days": port_kpi.get("NB_DAYS") or 0,
                    "max_drawdown": port_kpi.get("max_drawdown") or 0,
                },
            )
        return "ok"
    except Exception as e:
        Backtest.objects.filter(id=bt.id).update(status=Backtest.Status.FAILED, error_message=str(e))
        raise


@shared_task(bind=True)
def run_backtest_job_task(self, backtest_id: int, user_id=None, job_id=None):
    """Tracked job wrapper around run_backtest_task."""
    job = None
    if job_id:
        job = ProcessingJob.objects.filter(id=job_id).first()
        if job:
            job.status = ProcessingJob.Status.RUNNING
            job.task_id = getattr(self.request, "id", "") or ""
            job.started_at = timezone.now()
            job.save(update_fields=["status", "task_id", "started_at"])
    if job is None:
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.RUN_BACKTEST,
            status=ProcessingJob.Status.RUNNING,
            task_id=getattr(self.request, "id", "") or "",
            backtest_id=backtest_id,
            created_by_id=user_id,
            started_at=timezone.now(),
        )
    try:
        _job_checkpoint(job)
        msg = run_backtest_task(backtest_id)
        job.status = ProcessingJob.Status.DONE
        job.message = str(msg)
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return msg
    except JobCancelled:
        job.status = ProcessingJob.Status.CANCELLED
        job.message = (job.message or "") + "\nCancelled by user (before start)."
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return "cancelled"
    except JobKilled:
        job.status = ProcessingJob.Status.KILLED
        job.message = (job.message or "") + "\nKilled by user (before start)."
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "message", "finished_at"])
        return "killed"
    except Exception as e:
        job.status = ProcessingJob.Status.FAILED
        job.error = str(e)
        job.finished_at = timezone.now()
        job.save(update_fields=["status", "error", "finished_at"])
        raise


@shared_task
def cleanup_stale_processing_jobs_task() -> dict:
    """Mark stale / zombie ProcessingJob rows as FAILED to keep /jobs and admin usable.

    This is defensive: if a worker is killed, redeployed, or crashes, the DB row can stay RUNNING/PENDING forever.

    Rules (conservative):
    - RUNNING with heartbeat older than JOB_STALE_HEARTBEAT_MINUTES => FAILED
    - RUNNING with no heartbeat and started_at older than JOB_STALE_STARTED_MINUTES => FAILED
    - PENDING older than JOB_STALE_PENDING_MINUTES => FAILED (likely never picked up)
    """
    from django.conf import settings as dj_settings
    from django.db.models import Q
    from django.utils import timezone

    hb_min = int(getattr(dj_settings, "JOB_STALE_HEARTBEAT_MINUTES", 15))
    started_min = int(getattr(dj_settings, "JOB_STALE_STARTED_MINUTES", 30))
    pending_min = int(getattr(dj_settings, "JOB_STALE_PENDING_MINUTES", 60))

    now = timezone.now()
    from datetime import timedelta

    hb_cutoff = now - timedelta(minutes=hb_min)
    started_cutoff = now - timedelta(minutes=started_min)
    pending_cutoff = now - timedelta(minutes=pending_min)

    # RUNNING stale by heartbeat
    q_running_hb = Q(status=ProcessingJob.Status.RUNNING) & Q(heartbeat_at__isnull=False) & Q(heartbeat_at__lt=hb_cutoff)
    # RUNNING stale by started_at without heartbeat
    q_running_nohb = Q(status=ProcessingJob.Status.RUNNING) & Q(heartbeat_at__isnull=True) & Q(started_at__isnull=False) & Q(started_at__lt=started_cutoff)
    # PENDING too old
    q_pending_old = Q(status=ProcessingJob.Status.PENDING) & Q(created_at__lt=pending_cutoff)

    stale_qs = ProcessingJob.objects.filter(q_running_hb | q_running_nohb | q_pending_old)

    updated = 0
    for job in stale_qs.only("id", "status"):
        ProcessingJob.objects.filter(id=job.id).update(
            status=ProcessingJob.Status.FAILED,
            finished_at=now,
            error="Auto-marked as FAILED (stale job cleanup).",
        )
        updated += 1

    return {"updated": updated, "heartbeat_minutes": hb_min, "started_minutes": started_min, "pending_minutes": pending_min}


@shared_task(bind=True)
def export_scenario_xlsx_task(
    self,
    *,
    job_id: int,
    scenario_id: int,
    ticker: str = "",
    exchange: str = "",
    date_from: str = "",
    date_to: str = "",
):
    """Asynchronously export a Scenario workbook to /data/exports.

    Doing this in Celery avoids Gunicorn OOM/timeout on large scenario exports.
    """

    from pathlib import Path
    from django.utils import timezone
    from .models import Scenario, Symbol

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, "id", "") or ""
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.message = (
            f"Export XLSX scenario_id={scenario_id} ticker={ticker or '*'} exchange={exchange or '*'} "
            f"from={date_from or '-'} to={date_to or '-'}"
        )
        job.save(update_fields=["task_id", "status", "started_at", "message"])

    try:
        scenario = Scenario.objects.get(id=scenario_id)
        symbols_qs = Symbol.objects.filter(active=True)
        if ticker:
            symbols_qs = symbols_qs.filter(ticker=ticker)
        if exchange:
            symbols_qs = symbols_qs.filter(exchange=exchange)

        wb = build_scenario_workbook_write_only(
            scenario=scenario,
            symbols_qs=symbols_qs,
            date_from=date_from,
            date_to=date_to,
        )

        export_dir = Path("/data/exports/scenario") / str(scenario_id)
        export_dir.mkdir(parents=True, exist_ok=True)

        safe_name = "".join([c if c.isalnum() or c in ("-", "_") else "_" for c in (scenario.name or "scenario")])[:50] or "scenario"
        filename = f"{safe_name}_job{job_id}.xlsx"
        out_path = export_dir / filename

        wb.save(out_path)

        if job:
            job.status = ProcessingJob.Status.DONE
            job.finished_at = timezone.now()
            job.output_file = str(out_path)
            job.output_name = filename
            job.message = (job.message or "") + f"\n\nFichier prêt: {out_path}"
            job.save(update_fields=["status", "finished_at", "output_file", "output_name", "message"])

        return {"output": str(out_path)}

    except JobCancelled:
        if job:
            job.status = ProcessingJob.Status.CANCELLED
            job.finished_at = timezone.now()
            job.save(update_fields=["status", "finished_at"])
        return {"cancelled": True}
    except JobKilled:
        if job:
            job.status = ProcessingJob.Status.KILLED
            job.finished_at = timezone.now()
            job.save(update_fields=["status", "finished_at"])
        return {"killed": True}
    except Exception as e:
        if job:
            job.status = ProcessingJob.Status.FAILED
            job.finished_at = timezone.now()
            job.error = f"{type(e).__name__}: {e}"
            job.save(update_fields=["status", "finished_at", "error"])
        raise


def _exports_root() -> Path:
    root = Path('/data/exports')
    root.mkdir(parents=True, exist_ok=True)
    return root


def _job_export_path(job_id: int, filename: str) -> Path:
    safe_name = ''.join(ch if ch.isalnum() or ch in '._-' else '_' for ch in (filename or 'export.bin'))
    return _exports_root() / f'job_{job_id}_{safe_name}'


def _finalize_job_file(job: ProcessingJob | None, path: Path, output_name: str, message: str = '') -> str:
    if job:
        job.status = ProcessingJob.Status.DONE
        job.output_file = str(path)
        job.output_name = output_name
        job.message = message or job.message
        job.finished_at = timezone.now()
        job.save(update_fields=['status', 'output_file', 'output_name', 'message', 'finished_at'])
    return str(path)


def _fail_job(job: ProcessingJob | None, exc: Exception) -> None:
    if job:
        job.status = ProcessingJob.Status.FAILED
        job.error = str(exc)
        job.finished_at = timezone.now()
        job.save(update_fields=['status', 'error', 'finished_at'])


@shared_task(bind=True)
def export_alerts_csv_task(self, *, job_id: int, date_str: str = '', scenario_id: str = '', ticker: str = '', alert_codes: list[str] | None = None):
    from django.db.models import Q

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, 'id', '') or ''
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.save(update_fields=['task_id', 'status', 'started_at'])
    try:
        qs = Alert.objects.select_related('symbol', 'scenario').all().order_by('-date', 'scenario__name', 'symbol__ticker')
        if date_str:
            qs = qs.filter(date=date_str)
        if scenario_id:
            qs = qs.filter(scenario_id=scenario_id)
        if ticker:
            qs = qs.filter(symbol__ticker=ticker)
        if alert_codes:
            q = Q()
            for code in alert_codes:
                code = (code or '').strip()
                if code:
                    q |= Q(alerts__icontains=code)
            if q:
                qs = qs.filter(q)

        output_name = 'alerts_export.csv'
        path = _job_export_path(job_id, output_name)
        with path.open('w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['date', 'scenario_id', 'scenario_name', 'ticker', 'exchange', 'alerts'])
            for row in qs.iterator(chunk_size=2000):
                w.writerow([
                    row.date.isoformat() if row.date else '',
                    row.scenario_id or '',
                    getattr(row.scenario, 'name', '') if row.scenario_id else '',
                    getattr(row.symbol, 'ticker', '') if row.symbol_id else '',
                    getattr(row.symbol, 'exchange', '') if row.symbol_id else '',
                    row.alerts or '',
                ])
        return _finalize_job_file(job, path, output_name, f'Exported alerts CSV ({qs.count()} rows)')
    except Exception as exc:
        _fail_job(job, exc)
        raise


@shared_task(bind=True)
def export_all_scenarios_zip_task(self, *, job_id: int, ticker: str = '', exchange: str = '', date_from: str = '', date_to: str = ''):
    from .models import Scenario, Symbol
    from .views import _build_scenario_workbook

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, 'id', '') or ''
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.save(update_fields=['task_id', 'status', 'started_at'])
    try:
        scenarios = list(Scenario.objects.all().order_by('name', 'id'))
        symbols_qs = Symbol.objects.filter(active=True)
        if ticker:
            symbols_qs = symbols_qs.filter(ticker=ticker)
        if exchange:
            symbols_qs = symbols_qs.filter(exchange=exchange)

        output_name = 'all_scenarios_export.zip'
        path = _job_export_path(job_id, output_name)
        with zipfile.ZipFile(path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            for scenario in scenarios:
                wb = _build_scenario_workbook(scenario, symbols_qs, date_from=date_from, date_to=date_to)
                buf = io.BytesIO()
                wb.save(buf)
                safe_scn = ''.join(ch if ch.isalnum() or ch in '._-' else '_' for ch in scenario.name)
                zf.writestr(f'scenario_{scenario.id}_{safe_scn}.xlsx', buf.getvalue())
        return _finalize_job_file(job, path, output_name, f'Exported ZIP for {len(scenarios)} scenarios')
    except Exception as exc:
        _fail_job(job, exc)
        raise


@shared_task(bind=True)
def export_data_xlsx_task(self, *, job_id: int, ticker: str = '', exchange: str = '', scenario_id: str = '', date_from: str = '', date_to: str = ''):
    from openpyxl import Workbook
    from .models import Scenario, Symbol, DailyBar, DailyMetric, Alert

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, 'id', '') or ''
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.save(update_fields=['task_id', 'status', 'started_at'])
    try:
        wb = Workbook()
        ws_bars = wb.active
        ws_bars.title = 'Bars'
        ws_bars.append(['ticker', 'exchange', 'date', 'open', 'high', 'low', 'close', 'volume'])

        symbols_qs = Symbol.objects.filter(active=True)
        if ticker:
            symbols_qs = symbols_qs.filter(ticker=ticker)
        if exchange:
            symbols_qs = symbols_qs.filter(exchange=exchange)
        symbol_ids = list(symbols_qs.values_list('id', flat=True))

        bars = DailyBar.objects.filter(symbol_id__in=symbol_ids).select_related('symbol').order_by('symbol__ticker', 'date')
        if date_from:
            bars = bars.filter(date__gte=date_from)
        if date_to:
            bars = bars.filter(date__lte=date_to)
        for b in bars.iterator(chunk_size=2000):
            ws_bars.append([b.symbol.ticker, b.symbol.exchange, b.date.isoformat(), b.open, b.high, b.low, b.close, b.volume])

        ws_metrics = wb.create_sheet('Metrics')
        ws_metrics.append(['scenario_id', 'scenario_name', 'ticker', 'date', 'P', 'M', 'M1', 'X', 'X1', 'T', 'Q', 'S', 'K1', 'Kf', 'K2', 'K3', 'K4', 'sum_slope', 'slope_vrai', 'sum_slope_basse', 'slope_vrai_basse', 'ratio_P'])
        metrics = DailyMetric.objects.select_related('scenario', 'symbol').filter(symbol_id__in=symbol_ids).order_by('scenario__name', 'symbol__ticker', 'date')
        if scenario_id:
            metrics = metrics.filter(scenario_id=scenario_id)
        if date_from:
            metrics = metrics.filter(date__gte=date_from)
        if date_to:
            metrics = metrics.filter(date__lte=date_to)
        for m in metrics.iterator(chunk_size=2000):
            ws_metrics.append([m.scenario_id, m.scenario.name if m.scenario_id else '', m.symbol.ticker if m.symbol_id else '', m.date.isoformat(), m.P, m.M, m.M1, m.X, m.X1, m.T, m.Q, m.S, m.K1, getattr(m, 'Kf2bis', None), m.K2, m.K3, m.K4, m.sum_slope, m.slope_vrai, m.sum_slope_basse, m.slope_vrai_basse, m.ratio_P])

        ws_alerts = wb.create_sheet('Alerts')
        ws_alerts.append(['scenario_id', 'scenario_name', 'ticker', 'exchange', 'date', 'alerts'])
        alerts = Alert.objects.select_related('scenario', 'symbol').filter(symbol_id__in=symbol_ids).order_by('-date', 'scenario__name', 'symbol__ticker')
        if scenario_id:
            alerts = alerts.filter(scenario_id=scenario_id)
        if date_from:
            alerts = alerts.filter(date__gte=date_from)
        if date_to:
            alerts = alerts.filter(date__lte=date_to)
        for a in alerts.iterator(chunk_size=2000):
            ws_alerts.append([a.scenario_id, a.scenario.name if a.scenario_id else '', a.symbol.ticker if a.symbol_id else '', a.symbol.exchange if a.symbol_id else '', a.date.isoformat(), a.alerts or ''])

        output_name = 'data_export.xlsx'
        path = _job_export_path(job_id, output_name)
        wb.save(path)
        return _finalize_job_file(job, path, output_name, 'Exported combined data workbook')
    except Exception as exc:
        _fail_job(job, exc)
        raise


@shared_task(bind=True)
def export_backtest_debug_csv_task(self, *, job_id: int, backtest_id: int, ticker: str = '', line: str = ''):
    from .models import Backtest

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, 'id', '') or ''
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.save(update_fields=['task_id', 'status', 'started_at'])
    try:
        bt = Backtest.objects.get(id=backtest_id)
        results = bt.results or {}
        tickers_map = results.get('tickers') or {}
        if not ticker:
            ticker = next(iter(tickers_map.keys()), '')
        if not ticker or ticker not in tickers_map:
            raise ValueError('Ticker introuvable dans les résultats du backtest.')
        lines = (tickers_map.get(ticker) or {}).get('lines') or []
        if line != '':
            try:
                line_idx = int(line)
                selected = [ln for ln in lines if int(ln.get('line_index') or -1) == line_idx]
            except Exception:
                selected = []
        else:
            selected = lines[:1]
        if not selected:
            raise ValueError('Ligne introuvable dans les résultats du backtest.')
        daily = selected[0].get('daily') or []
        output_name = f'backtest_{backtest_id}_{ticker}_debug.csv'
        path = _job_export_path(job_id, output_name)
        fieldnames = sorted({k for row in daily for k in row.keys()})
        with path.open('w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for row in daily:
                w.writerow(row)
        return _finalize_job_file(job, path, output_name, f'Exported backtest debug CSV for {ticker}')
    except Exception as exc:
        _fail_job(job, exc)
        raise


@shared_task(bind=True)
def export_backtest_excel_task(self, *, job_id: int, backtest_id: int):
    from .models import Backtest
    from .views import _build_backtest_workbook_full

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, 'id', '') or ''
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.save(update_fields=['task_id', 'status', 'started_at'])
    try:
        bt = Backtest.objects.get(id=backtest_id)
        wb, base_name = _build_backtest_workbook_full(bt)
        output_name = f'{base_name}.xlsx'
        path = _job_export_path(job_id, output_name)
        wb.save(path)
        return _finalize_job_file(job, path, output_name, f'Exported full backtest workbook for #{backtest_id}')
    except Exception as exc:
        _fail_job(job, exc)
        raise


@shared_task(bind=True)
def export_backtest_excel_compact_task(self, *, job_id: int, backtest_id: int, charts: str = '1', chart_mode: str = 'top', chart_limit: str = '', chart_ticker: str = '', chart_line: str = ''):
    from .models import Backtest
    from .views import _build_backtest_workbook_compact

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, 'id', '') or ''
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.save(update_fields=['task_id', 'status', 'started_at'])
    try:
        bt = Backtest.objects.get(id=backtest_id)
        wb, base_name = _build_backtest_workbook_compact(bt, charts=charts, chart_mode=chart_mode, chart_limit=chart_limit, chart_ticker=chart_ticker, chart_line=chart_line)
        output_name = f'{base_name}.xlsx'
        path = _job_export_path(job_id, output_name)
        wb.save(path)
        return _finalize_job_file(job, path, output_name, f'Exported compact backtest workbook for #{backtest_id}')
    except Exception as exc:
        _fail_job(job, exc)
        raise


@shared_task(bind=True)
def export_game_scenario_xlsx_task(self, *, job_id: int, game_scenario_id: int):
    from openpyxl import Workbook
    from .models import GameScenario

    job = ProcessingJob.objects.filter(id=job_id).first()
    if job:
        job.task_id = getattr(self.request, 'id', '') or ''
        job.status = ProcessingJob.Status.RUNNING
        job.started_at = timezone.now()
        job.save(update_fields=['task_id', 'status', 'started_at'])
    try:
        game = GameScenario.objects.get(id=game_scenario_id)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Game'
        ws.append(['Field', 'Value'])
        for key, value in [
            ('id', game.id),
            ('name', game.name),
            ('description', game.description or ''),
            ('active', game.active),
            ('study_days', game.study_days),
            ('capital_total', game.capital_total),
            ('capital_per_ticker', game.capital_per_ticker),
            ('tradability_threshold', game.tradability_threshold),
            ('presence_threshold_pct', getattr(game, 'presence_threshold_pct', '')),
            ('warmup_days', getattr(game, 'warmup_days', '')),
            ('last_run_at', game.last_run_at.isoformat() if game.last_run_at else ''),
            ('last_run_status', game.last_run_status or ''),
        ]:
            ws.append([key, value])
        ws2 = wb.create_sheet('TodayResults')
        rows = game.today_results or []
        headers = sorted({k for row in rows for k in row.keys()}) if rows else ['ticker', 'best_bmd', 'ok']
        ws2.append(headers)
        for row in rows:
            ws2.append([row.get(h) for h in headers])
        output_name = f'game_scenario_{game_scenario_id}.xlsx'
        path = _job_export_path(job_id, output_name)
        wb.save(path)
        return _finalize_job_file(job, path, output_name, f'Exported GameScenario workbook for #{game_scenario_id}')
    except Exception as exc:
        _fail_job(job, exc)
        raise
