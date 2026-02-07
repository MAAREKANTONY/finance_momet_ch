import csv
import os
from io import BytesIO
from typing import Iterable
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Max, Q
from django.db.models.deletion import ProtectedError
from django.views.decorators.http import require_POST

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, Reference
import zipfile as pyzip
import json
import tempfile
from pathlib import Path

from .models import Alert, Scenario, Symbol, EmailRecipient, DailyBar, DailyMetric, EmailSettings, JobLog, Backtest, ProcessingJob
from .forms import ScenarioForm, EmailRecipientForm, SymbolManualForm, EmailSettingsForm, SymbolScenariosForm, SymbolImportForm, BacktestForm, BACKTEST_SIGNAL_CHOICES
from .services.provider_twelvedata import TwelveDataClient
from .services.backtesting.parquet_storage import parquet_storage_enabled
from .services.backtesting.volume_guards import should_limit_excel, select_top_tickers_by_metric, excel_full_tickers_threshold, excel_top_n

try:
    # Celery is optional in dev; we keep the import defensive so the web container can boot.
    from .tasks import fetch_daily_bars_task
except Exception:  # pragma: no cover
    fetch_daily_bars_task = None


def _refresh_backtest_universe_snapshot(bt: Backtest) -> None:
    """Refresh the backtest universe_snapshot from the current Scenario symbols.

    Why:
    - Users expect changes to a Scenario universe (tickers added/removed) to be taken into
      account when they re-run Fetch/Compute/Backtest.
    - We keep the snapshot concept (it is still stored on the Backtest), but we update it
      whenever the user explicitly triggers a new processing action.
    """
    symbols = (
        bt.scenario.symbols.all()
        .order_by("ticker", "exchange")
        .values_list("ticker", "exchange")
    )
    bt.universe_snapshot = [{"ticker": t, "exchange": e} for t, e in symbols]
    bt.save(update_fields=["universe_snapshot", "updated_at"])


@login_required
def dashboard(request):
    last = Alert.objects.aggregate(last_date=Max("date"))["last_date"]
    scenarios = Scenario.objects.filter(active=True).count()
    symbols = Symbol.objects.filter(active=True).count()
    alerts_count = Alert.objects.filter(date=last).count() if last else 0
    return render(request, "dashboard.html", {"last_date": last, "scenarios": scenarios, "symbols": symbols, "alerts_count": alerts_count})



@login_required
def alerts_table(request):
    date_str = (request.GET.get("date") or "").strip()
    scenario_id = (request.GET.get("scenario") or "").strip()
    ticker = (request.GET.get("ticker") or "").strip()
    alert_codes = request.GET.getlist("alert")

    qs = Alert.objects.select_related("symbol", "scenario").all().order_by("-date", "scenario__name", "symbol__ticker")
    if date_str:
        qs = qs.filter(date=date_str)
    if scenario_id:
        qs = qs.filter(scenario_id=scenario_id)
    if ticker:
        qs = qs.filter(symbol__ticker=ticker)
    if alert_codes:
        from django.db.models import Q
        q = Q()
        for code in alert_codes:
            if code:
                q |= Q(alerts__icontains=code)
        qs = qs.filter(q)

    scenarios = Scenario.objects.all().order_by("name")
    symbols = Symbol.objects.all().order_by("ticker")
    # Keep in sync with calculations.py (crossing rules)
    all_alert_codes = ["A1","B1","A1f","B1f","C1","D1","E1","F1","G1","H1"]

    try:
        from core.services.backtesting.volume_guards import volume_guards_enabled
        guards_on = volume_guards_enabled()
    except Exception:
        guards_on = False

    return render(request, "alerts.html", {
        "alerts": qs[:2000],
        "scenarios": scenarios,
        "symbols": symbols,
        "selected_date": date_str,
        "selected_scenario": int(scenario_id) if scenario_id else "",
        "selected_ticker": ticker,
        "selected_alerts": alert_codes,
        "all_alert_codes": all_alert_codes,
        "ENABLE_VOLUME_GUARDS": guards_on,
    })


@login_required
def alerts_export_csv(request):
    date_str = request.GET.get("date") or ""
    scenario_id = request.GET.get("scenario") or ""
    qs = Alert.objects.select_related("symbol", "scenario").all().order_by("date", "scenario__name", "symbol__ticker")
    if date_str:
        qs = qs.filter(date=date_str)
    if scenario_id:
        qs = qs.filter(scenario_id=scenario_id)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="alerts.csv"'
    writer = csv.writer(response)
    writer.writerow(["date", "ticker", "exchange", "scenario", "alerts"])
    for a in qs:
        writer.writerow([a.date.isoformat(), a.symbol.ticker, a.symbol.exchange, a.scenario.name, a.alerts])
    return response


def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(70, max(10, max_len + 2))


def _header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _build_scenario_workbook(scenario: Scenario, symbols_qs, date_from: str = "", date_to: str = "") -> Workbook:
    wb = Workbook()
    first = True

    alerts_qs = Alert.objects.filter(scenario=scenario, symbol__in=symbols_qs).select_related("symbol").order_by("symbol__ticker", "date")
    if date_from:
        alerts_qs = alerts_qs.filter(date__gte=date_from)
    if date_to:
        alerts_qs = alerts_qs.filter(date__lte=date_to)
    alerts_map = {(a.symbol_id, a.date): a.alerts for a in alerts_qs}

    for sym in symbols_qs.order_by("ticker", "exchange"):
        bars = DailyBar.objects.filter(symbol=sym).order_by("date")
        if date_from:
            bars = bars.filter(date__gte=date_from)
        if date_to:
            bars = bars.filter(date__lte=date_to)

        metrics = DailyMetric.objects.filter(symbol=sym, scenario=scenario).order_by("date")
        if date_from:
            metrics = metrics.filter(date__gte=date_from)
        if date_to:
            metrics = metrics.filter(date__lte=date_to)
        metrics_by_date = {m.date: m for m in metrics}

        title = sym.ticker[:28]
        ws = wb.active if first else wb.create_sheet(title=title)
        ws.title = title
        first = False

        ws.append([f"Scenario: {scenario.name}"])
        ws.append([f"Description: {scenario.description}"])
        ws.append([f"Vars: a={scenario.a} b={scenario.b} c={scenario.c} d={scenario.d} e={scenario.e} vc={getattr(scenario,'vc',None)} fl={getattr(scenario,'fl',None)} | N1={scenario.n1} N2={scenario.n2} N3={scenario.n3} | history_years={scenario.history_years}"])
        ws.append([f"Ticker: {sym.ticker}  Exchange: {sym.exchange}  Name: {sym.name}"])
        ws.append([])

        header = [
            "date",
            "open","high","low","close","volume","change_amount","change_pct",
            "V","slope_P","sum_pos_P","nb_pos_P","ratio_P","amp_h",
            "P","M","M1","X","X1","T","Q","S","K1","K1f","K2","K3","K4",
            "alerts",
        ]
        ws.append(header)
        _header(ws, ws.max_row)

        def f(x):
            return float(x) if x is not None else None

        for b in bars:
            m = metrics_by_date.get(b.date)
            ws.append([
                b.date.isoformat(),
                f(b.open), f(b.high), f(b.low), f(b.close), (int(b.volume) if b.volume is not None else None), f(b.change_amount), f(b.change_pct),
                f(m.V) if m else None,
                f(m.slope_P) if m else None,
                f(m.sum_pos_P) if m else None,
                (m.nb_pos_P if m and m.nb_pos_P is not None else None),
                f(m.ratio_P) if m else None,
                f(m.amp_h) if m else None,
                f(m.P) if m else None,
                f(m.M) if m else None,
                f(m.M1) if m else None,
                f(m.X) if m else None,
                f(m.X1) if m else None,
                f(m.T) if m else None,
                f(m.Q) if m else None,
                f(m.S) if m else None,
                f(m.K1) if m else None,
                f(m.K1f) if m else None,
                f(m.K2) if m else None,
                f(m.K3) if m else None,
                f(m.K4) if m else None,
                alerts_map.get((sym.id, b.date), ""),
            ])

        _autosize(ws)

    return wb


def _alerts_excel_select_symbols_with_guards(symbols_qs, scenario: Scenario, date_from: str = "", date_to: str = ""):
    """Apply Excel volume guards for Alerts scenario exports.

    When ENABLE_VOLUME_GUARDS=1 and the number of tickers exceeds EXCEL_FULL_TICKERS_THRESHOLD,
    we only export Top N tickers (deterministic order by ticker) and include a Summary sheet.

    This function is additive and does not affect legacy exports.
    """
    try:
        from core.services.backtesting.volume_guards import should_limit_excel, excel_top_n
    except Exception:
        return symbols_qs, None

    total = symbols_qs.count()
    if not should_limit_excel(total):
        return symbols_qs, None

    top_n = max(1, int(excel_top_n()))
    selected = symbols_qs.order_by("ticker", "exchange")[:top_n]
    info = {
        "limited": True,
        "total": total,
        "exported": selected.count(),
        "top_n": top_n,
        "scenario": scenario.name,
        "date_from": date_from,
        "date_to": date_to,
    }
    return selected, info


def _alerts_add_summary_sheet(wb: Workbook, info: dict, all_symbols_qs):
    """Add a lightweight Summary sheet when Excel guards are applied."""
    if not info or not info.get("limited"):
        return
    ws = wb.create_sheet(title="Summary", index=0)
    ws.append(["EXPORT LIMIT ACTIVATED"])
    ws.append(["Reason", "Too many tickers for Excel export"])
    ws.append(["Scenario", info.get("scenario")])
    ws.append(["Total tickers", info.get("total")])
    ws.append(["Exported tickers", info.get("exported")])
    ws.append(["Top N", info.get("top_n")])
    ws.append(["Date from", info.get("date_from") or "(none)"])
    ws.append(["Date to", info.get("date_to") or "(none)"])
    ws.append([])
    ws.append(["Note"])
    ws.append([
        "Use 'Export Détails (Parquet/CSV ZIP)' to retrieve the complete dataset for all tickers."
    ])
    ws.append([])
    ws.append(["All active tickers (deterministic order)"])
    ws.append(["ticker", "exchange"])
    _header(ws, ws.max_row)
    for s in all_symbols_qs.order_by("ticker", "exchange"):
        ws.append([s.ticker, s.exchange])
    _autosize(ws)


@login_required
def data_export_scenario_xlsx(request, scenario_id: int):
    scenario = get_object_or_404(Scenario, pk=scenario_id)
    ticker = (request.GET.get("ticker") or "").strip()
    exchange = (request.GET.get("exchange") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    symbols_qs = Symbol.objects.filter(active=True)
    if ticker:
        symbols_qs = symbols_qs.filter(ticker=ticker)
    if exchange:
        symbols_qs = symbols_qs.filter(exchange=exchange)

    # Volume guards (optional): limit number of tickers in Excel exports to avoid memory blowups.
    all_symbols_qs = symbols_qs
    symbols_qs, guard_info = _alerts_excel_select_symbols_with_guards(
        symbols_qs, scenario=scenario, date_from=date_from, date_to=date_to
    )

    wb = _build_scenario_workbook(scenario, symbols_qs, date_from=date_from, date_to=date_to)
    if guard_info:
        _alerts_add_summary_sheet(wb, guard_info, all_symbols_qs)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_name = "".join([c if c.isalnum() or c in ("-","_") else "_" for c in scenario.name])[:50] or "scenario"
    resp = HttpResponse(bio.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{safe_name}.xlsx"'
    return resp


@login_required
def data_export_all_scenarios_zip(request):
    ticker = (request.GET.get("ticker") or "").strip()
    exchange = (request.GET.get("exchange") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    symbols_qs = Symbol.objects.filter(active=True)
    if ticker:
        symbols_qs = symbols_qs.filter(ticker=ticker)
    if exchange:
        symbols_qs = symbols_qs.filter(exchange=exchange)

    scenarios = Scenario.objects.filter(active=True).order_by("name")

    # Create ZIP on disk to avoid memory blowups for large exports
    tmp = tempfile.NamedTemporaryFile(prefix="alerts_scenarios_", suffix=".zip", delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()

    with pyzip.ZipFile(tmp_path, "w", compression=pyzip.ZIP_DEFLATED) as zf:
        for scenario in scenarios:
            all_symbols_qs = symbols_qs
            scoped_symbols_qs, guard_info = _alerts_excel_select_symbols_with_guards(
                symbols_qs, scenario=scenario, date_from=date_from, date_to=date_to
            )
            wb = _build_scenario_workbook(scenario, scoped_symbols_qs, date_from=date_from, date_to=date_to)
            if guard_info:
                _alerts_add_summary_sheet(wb, guard_info, all_symbols_qs)
            wb_io = BytesIO(); wb.save(wb_io); wb_io.seek(0)
            safe_name = "".join([c if c.isalnum() or c in ("-","_") else "_" for c in scenario.name])[:50] or "scenario"
            zf.writestr(f"{safe_name}.xlsx", wb_io.getvalue())

    tmp_f = open(tmp_path, "rb")
    resp = FileResponse(tmp_f, as_attachment=True, filename="scenarios_exports.zip", content_type="application/zip")
    return resp
def _build_alerts_details_table_for_symbol(sym: Symbol, scenario: Scenario, date_from: str = "", date_to: str = ""):
    """Build a per-ticker pyarrow Table with Bars + Metrics + Alerts (scenario-scoped).

    This mirrors the per-ticker sheet structure from the scenario Excel export but is optimized
    for large volumes.
    """
    # Late imports (optional dependency)
    import pyarrow as pa  # type: ignore

    bars = DailyBar.objects.filter(symbol=sym).order_by("date")
    if date_from:
        bars = bars.filter(date__gte=date_from)
    if date_to:
        bars = bars.filter(date__lte=date_to)

    metrics = DailyMetric.objects.filter(symbol=sym, scenario=scenario).order_by("date")
    if date_from:
        metrics = metrics.filter(date__gte=date_from)
    if date_to:
        metrics = metrics.filter(date__lte=date_to)
    metrics_by_date = {m.date: m for m in metrics}

    alerts = Alert.objects.filter(symbol=sym, scenario=scenario).order_by("date")
    if date_from:
        alerts = alerts.filter(date__gte=date_from)
    if date_to:
        alerts = alerts.filter(date__lte=date_to)
    alerts_by_date = {a.date: a.alerts for a in alerts}

    rows = []

    def f(x):
        try:
            return float(x) if x is not None else None
        except Exception:
            return None

    for b in bars:
        m = metrics_by_date.get(b.date)
        rows.append({
            "date": b.date.isoformat(),
            "open": f(b.open), "high": f(b.high), "low": f(b.low), "close": f(b.close),
            "volume": int(b.volume) if getattr(b, "volume", None) is not None else None,
            "change_amount": f(b.change_amount),
            "change_pct": f(b.change_pct),
            "V": f(m.V) if m else None,
            "slope_P": f(m.slope_P) if m else None,
            "sum_pos_P": f(m.sum_pos_P) if m else None,
            "nb_pos_P": int(m.nb_pos_P) if (m and m.nb_pos_P is not None) else None,
            "ratio_P": f(m.ratio_P) if m else None,
            "amp_h": f(m.amp_h) if m else None,
            "P": f(m.P) if m else None,
            "M": f(m.M) if m else None,
            "M1": f(m.M1) if m else None,
            "X": f(m.X) if m else None,
            "X1": f(m.X1) if m else None,
            "T": f(m.T) if m else None,
            "Q": f(m.Q) if m else None,
            "S": f(m.S) if m else None,
            "K1": f(m.K1) if m else None,
            "K1f": f(m.K1f) if m else None,
            "K2": f(m.K2) if m else None,
            "K3": f(m.K3) if m else None,
            "K4": f(m.K4) if m else None,
            "alerts": alerts_by_date.get(b.date, ""),
        })

    return pa.Table.from_pylist(rows)


@login_required
def data_export_scenario_details_zip(request, scenario_id: int):
    """Optimized per-scenario export: ZIP of Parquet (or CSV) files, 1 per ticker.

    This is designed for high-volume use cases (hundreds of tickers) and avoids Excel memory limits.
    """
    scenario = get_object_or_404(Scenario, pk=scenario_id)
    ticker = (request.GET.get("ticker") or "").strip()
    exchange = (request.GET.get("exchange") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()
    fmt = (request.GET.get("format") or "parquet").strip().lower()
    if fmt not in {"parquet", "csv"}:
        fmt = "parquet"

    symbols_qs = Symbol.objects.filter(active=True)
    if ticker:
        symbols_qs = symbols_qs.filter(ticker=ticker)
    if exchange:
        symbols_qs = symbols_qs.filter(exchange=exchange)

    # Late imports (optional dependency)
    try:
        import pyarrow.parquet as pq  # type: ignore
        import pyarrow.csv as pacsv  # type: ignore
    except Exception as e:
        messages.error(request, f"pyarrow requis pour l'export détails: {e}")
        return redirect("alerts_table")

    tmp = tempfile.NamedTemporaryFile(prefix=f"alerts_scenario_{scenario.id}_details_", suffix=".zip", delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()

    safe_name = "".join([c if c.isalnum() or c in ("-","_") else "_" for c in scenario.name])[:50] or "scenario"
    try:
        with pyzip.ZipFile(tmp_path, "w", compression=pyzip.ZIP_DEFLATED) as zf:
            for sym in symbols_qs.order_by("ticker", "exchange"):
                table = _build_alerts_details_table_for_symbol(sym, scenario, date_from=date_from, date_to=date_to)
                if fmt == "parquet":
                    buf = BytesIO()
                    pq.write_table(table, buf)
                    buf.seek(0)
                    zf.writestr(f"{sym.ticker}.parquet", buf.getvalue())
                else:
                    table = _arrow_table_to_csv_safe(table)
                    buf = BytesIO(); pacsv.write_csv(table, buf); buf.seek(0)
                    zf.writestr(f"{sym.ticker}.csv", buf.getvalue())

        tmp_f = open(tmp_path, "rb")
        filename = f"alerts_{safe_name}_details_{fmt}.zip"
        return FileResponse(tmp_f, as_attachment=True, filename=filename, content_type="application/zip")
    finally:
        pass


@login_required
def data_export_xlsx(request):
    # Legacy combined export retained (3 sheets)
    ticker = (request.GET.get("ticker") or "").strip()
    exchange = (request.GET.get("exchange") or "").strip()
    scenario_id = (request.GET.get("scenario") or "").strip()
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    symbols_qs = Symbol.objects.all()
    if ticker:
        symbols_qs = symbols_qs.filter(ticker=ticker)
    if exchange:
        symbols_qs = symbols_qs.filter(exchange=exchange)

    symbol_ids = list(symbols_qs.values_list("id", flat=True))
    bars = DailyBar.objects.filter(symbol_id__in=symbol_ids).select_related("symbol").order_by("symbol__ticker", "date")
    if date_from:
        bars = bars.filter(date__gte=date_from)
    if date_to:
        bars = bars.filter(date__lte=date_to)

    metrics = DailyMetric.objects.filter(symbol_id__in=symbol_ids).select_related("symbol", "scenario").order_by("symbol__ticker", "scenario__name", "date")
    if scenario_id:
        metrics = metrics.filter(scenario_id=scenario_id)
    if date_from:
        metrics = metrics.filter(date__gte=date_from)
    if date_to:
        metrics = metrics.filter(date__lte=date_to)

    alerts = Alert.objects.filter(symbol_id__in=symbol_ids).select_related("symbol", "scenario").order_by("symbol__ticker", "scenario__name", "date")
    if scenario_id:
        alerts = alerts.filter(scenario_id=scenario_id)
    if date_from:
        alerts = alerts.filter(date__gte=date_from)
    if date_to:
        alerts = alerts.filter(date__lte=date_to)

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("DailyBars")
    header1 = ["ticker", "exchange", "date", "open", "high", "low", "close", "volume", "change_amount", "change_pct", "source"]
    ws1.append(header1); _header(ws1, 1)
    for b in bars:
        ws1.append([b.symbol.ticker, b.symbol.exchange, b.date.isoformat(), float(b.open), float(b.high), float(b.low), float(b.close),
                    int(b.volume) if getattr(b, "volume", None) is not None else None,
                    float(b.change_amount) if b.change_amount is not None else None,
                    float(b.change_pct) if b.change_pct is not None else None, b.source])
    _autosize(ws1)

    ws2 = wb.create_sheet("DailyMetrics")
    header2 = ["ticker","exchange","scenario","date","P","M","M1","X","X1","T","Q","S","K1","K2","K3","K4"]
    ws2.append(header2); _header(ws2, 1)
    def f(x): return float(x) if x is not None else None
    for m in metrics:
        ws2.append([m.symbol.ticker, m.symbol.exchange, m.scenario.name, m.date.isoformat(), f(m.P), f(m.M), f(m.M1), f(m.X), f(m.X1), f(m.T), f(m.Q), f(m.S), f(m.K1), f(m.K2), f(m.K3), f(m.K4)])
    _autosize(ws2)

    ws3 = wb.create_sheet("Alerts")
    header3 = ["date","ticker","exchange","scenario","alerts"]
    ws3.append(header3); _header(ws3, 1)
    for a in alerts:
        ws3.append([a.date.isoformat(), a.symbol.ticker, a.symbol.exchange, a.scenario.name, a.alerts])
    _autosize(ws3)

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="stock_data_export.xlsx"'
    return resp


@login_required
def symbols_page(request):
    symbols = Symbol.objects.all().order_by("-active", "ticker")[:2000]
    manual_form = SymbolManualForm()
    return render(request, "symbols.html", {"symbols": symbols, "manual_form": manual_form})


@login_required
@require_POST
def symbol_add(request):
    ticker = (request.POST.get("ticker") or "").strip()
    exchange = (request.POST.get("exchange") or "").strip()
    name = (request.POST.get("name") or "").strip()
    instrument_type = (request.POST.get("instrument_type") or "").strip()
    country = (request.POST.get("country") or "").strip()
    currency = (request.POST.get("currency") or "").strip()

    default_scenario = Scenario.objects.filter(is_default=True, active=True).first()

    if ticker:
        obj, created = Symbol.objects.get_or_create(
            ticker=ticker,
            exchange=exchange,
            defaults={
                "name": name,
                "instrument_type": instrument_type,
                "country": country,
                "currency": currency,
                "active": True,
            },
        )
        if not created:
            Symbol.objects.filter(id=obj.id).update(
                name=name or obj.name,
                instrument_type=instrument_type or obj.instrument_type,
                country=country or obj.country,
                currency=currency or obj.currency,
                active=True,
            )

        # Toujours associer au scénario par défaut si présent
        if default_scenario:
            obj.scenarios.add(default_scenario)

        messages.success(request, f"Ajouté: {ticker} {('('+exchange+')') if exchange else ''}")
        return redirect("symbols_page")

    form = SymbolManualForm(request.POST)
    if form.is_valid():
        sym = form.save()
        chosen = list(form.cleaned_data.get("scenarios") or [])
        if default_scenario and default_scenario not in chosen:
            chosen.append(default_scenario)
        if chosen:
            sym.scenarios.set(chosen)
        messages.success(request, f"Ajouté: {sym}")
    else:
        messages.error(request, "Erreur: symbole invalide.")
    return redirect("symbols_page")


@login_required
@require_POST
def symbol_toggle_active(request, pk: int):
    sym = get_object_or_404(Symbol, pk=pk)
    sym.active = not sym.active
    sym.save(update_fields=["active"])
    return redirect("symbols_page")


@login_required
@require_POST
def symbol_delete(request, pk: int):
    sym = get_object_or_404(Symbol, pk=pk)
    sym.delete()
    return redirect("symbols_page")


@login_required
def symbol_scenarios_edit(request, pk: int):
    """Manage scenario links from the ticker side."""

    symbol = get_object_or_404(Symbol, pk=pk)
    default_scenario = Scenario.objects.filter(is_default=True, active=True).first()

    if request.method == "POST":
        form = SymbolScenariosForm(request.POST)
        if form.is_valid():
            selected = list(form.cleaned_data["scenarios"])
            # Always keep default scenario linked if it exists
            if default_scenario and default_scenario not in selected:
                selected.append(default_scenario)
            symbol.scenarios.set(selected)
            messages.success(request, "Scénarios mis à jour.")
            JobLog.objects.create(job="symbol_scenarios", level="INFO", message=f"Updated scenarios for {symbol}")
            return redirect("symbols_page")
        messages.error(request, "Formulaire invalide.")
    else:
        initial = symbol.scenarios.filter(active=True)
        form = SymbolScenariosForm(initial={"scenarios": initial})

    return render(
        request,
        "symbol_scenarios.html",
        {"symbol": symbol, "form": form, "default_scenario": default_scenario},
    )


def _iter_symbol_rows_from_csv(file_obj) -> Iterable[dict]:
    """Yield dict rows from a CSV.

    Supports:
      - header-based CSV (DictReader)
      - headerless CSV with columns: ticker, exchange, scenarios
    """

    content = file_obj.read()
    if isinstance(content, bytes):
        # try utf-8 first, fallback to latin-1
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
    else:
        text = str(content)

    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return

    sample = "\n".join(lines[:50])[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel

    # First pass: read first row to decide if it looks like headers
    reader0 = csv.reader(lines, dialect=dialect)
    first_row = next(reader0, [])
    norm = [str(c).strip().lower() for c in first_row]

    header_keywords = {
        "ticker", "code", "ticker code", "ticker_code",
        "exchange", "market", "ticker market", "ticker_market",
        "scenario", "scenarios", "scenario list", "scenario_list",
    }
    looks_like_header = any(any(k in cell for k in header_keywords) for cell in norm)

    if looks_like_header:
        reader = csv.DictReader(lines, dialect=dialect)
        for row in reader:
            yield {
                (k.strip() if isinstance(k, str) else k): (v.strip() if isinstance(v, str) else v)
                for k, v in (row or {}).items()
            }
    else:
        # headerless mode: map columns by position
        # col0=ticker, col1=exchange, col2=scenarios
        reader = csv.reader(lines, dialect=dialect)
        for values in reader:
            if not values:
                continue
            d = {}
            if len(values) >= 1:
                d["ticker"] = str(values[0]).strip()
            if len(values) >= 2:
                d["exchange"] = str(values[1]).strip()
            if len(values) >= 3:
                d["scenario list"] = str(values[2]).strip()
            yield d

    content = file_obj.read()
    if isinstance(content, bytes):
        # try utf-8 first, fallback to latin-1
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
    else:
        text = str(content)

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    for row in reader:
        yield {k.strip() if isinstance(k, str) else k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}


def _iter_symbol_rows_from_xlsx(file_obj) -> Iterable[dict]:
    """Yield dict rows from an Excel file.

    Supports:
      - header row
      - headerless (assumes first columns: ticker, exchange, scenarios)
    """
    wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    first = next(rows, None)
    if not first:
        return

    first_norm = [str(h).strip().lower() if h is not None else "" for h in first]
    header_keywords = (
        "ticker", "code", "exchange", "market", "scenario"
    )
    looks_like_header = any(any(k in cell for k in header_keywords) for cell in first_norm)

    if looks_like_header:
        headers = [str(h).strip() if h is not None else "" for h in first]
        for values in rows:
            d = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = values[i] if i < len(values) else None
                if isinstance(v, str):
                    v = v.strip()
                d[h] = v
            yield d
    else:
        # headerless: first row is data
        def emit(values):
            d = {}
            if len(values) >= 1 and values[0] is not None:
                d["ticker"] = str(values[0]).strip()
            if len(values) >= 2 and values[1] is not None:
                d["exchange"] = str(values[1]).strip()
            if len(values) >= 3 and values[2] is not None:
                d["scenario list"] = str(values[2]).strip()
            return d

        yield emit(first)
        for values in rows:
            if values:
                yield emit(values)

@login_required
def symbols_import(request):
    """Bulk import tickers from CSV/XLSX.

    Expected columns (case-insensitive):
      - ticker code (MSFT)
      - ticker market (NASDAQ)
      - scenario list (scenario1, scenario2)
    """

    if request.method == "POST":
        form = SymbolImportForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data["file"]
            filename = (getattr(f, "name", "") or "").lower()
            default_scenario = Scenario.objects.filter(is_default=True, active=True).first()

            created = updated = skipped = 0
            missing_scenarios = 0
            errors: list[str] = []

            try:
                if filename.endswith(".xlsx") or filename.endswith(".xlsm") or filename.endswith(".xltx"):
                    row_iter = _iter_symbol_rows_from_xlsx(f)
                else:
                    row_iter = _iter_symbol_rows_from_csv(f)
            except Exception as e:
                messages.error(request, f"Impossible de lire le fichier: {e}")
                JobLog.objects.create(job="import_symbols", level="ERROR", message="Import failed", traceback=str(e))
                return redirect("symbols_page")

            def _get(row: dict, *keys: str) -> str:
                for k in keys:
                    for rk, rv in row.items():
                        if str(rk).strip().lower() == k:
                            return "" if rv is None else str(rv).strip()
                return ""

            for idx, row in enumerate(row_iter, start=2):
                ticker = _get(row, "ticker code", "ticker", "code", "ticker_code")
                market = _get(row, "ticker market", "market", "exchange", "ticker_market")
                scen_list = _get(row, "scenario list", "scenarios", "scenario", "scenario_list")

                if not ticker:
                    skipped += 1
                    continue

                try:
                    sym, was_created = Symbol.objects.get_or_create(
                        ticker=ticker,
                        exchange=market,
                        defaults={"active": True},
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                        if not sym.active:
                            sym.active = True
                            sym.save(update_fields=["active"])

                    selected_scenarios: list[Scenario] = []
                    if default_scenario:
                        selected_scenarios.append(default_scenario)

                    if scen_list:
                        for name in [s.strip() for s in scen_list.split(",") if s.strip()]:
                            scen = Scenario.objects.filter(name__iexact=name).first()
                            if scen and scen.active:
                                if scen not in selected_scenarios:
                                    selected_scenarios.append(scen)
                            else:
                                missing_scenarios += 1

                    if selected_scenarios:
                        sym.scenarios.add(*selected_scenarios)
                except Exception as e:
                    skipped += 1
                    msg = f"Ligne {idx}: erreur pour ticker={ticker} market={market}: {e}"
                    errors.append(msg)

            summary = (
                f"Import tickers terminé. created={created}, updated={updated}, skipped={skipped}, "
                f"scenario_not_found={missing_scenarios}."
            )
            details = "\n".join(errors[:80])
            JobLog.objects.create(
                job="import_symbols",
                level="ERROR" if errors else "INFO",
                message=summary + ("\n" + details if details else ""),
            )
            messages.success(request, summary)
            return redirect("symbols_page")
    else:
        form = SymbolImportForm()

    return render(request, "symbols_import.html", {"form": form})


@login_required
def scenarios_page(request):
    scenarios = Scenario.objects.all().order_by("-active", "name")
    return render(request, "scenarios.html", {"scenarios": scenarios})


@login_required
def scenario_create(request):
    has_other_default = Scenario.objects.filter(is_default=True).exists()
    if request.method == "POST":
        form = ScenarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Scénario créé.")
            return redirect("scenarios_page")
    else:
        form = ScenarioForm()
    return render(request, "scenario_form.html", {"form": form, "mode": "create", "has_other_default": has_other_default})


@login_required
def scenario_edit(request, pk: int):
    scenario = get_object_or_404(Scenario, pk=pk)
    has_other_default = Scenario.objects.filter(is_default=True).exclude(pk=scenario.pk).exists()
    if request.method == "POST":
        form = ScenarioForm(request.POST, instance=scenario)
        if form.is_valid():
            form.save()
            messages.success(request, "Scénario mis à jour.")
            return redirect("scenarios_page")
    else:
        form = ScenarioForm(instance=scenario)
    return render(request, "scenario_form.html", {"form": form, "mode": "edit", "scenario": scenario, "has_other_default": has_other_default})


@login_required
@require_POST
def scenario_delete(request, pk: int):
    scenario = get_object_or_404(Scenario, pk=pk)
    try:
        scenario.delete()
        messages.success(request, "Scénario supprimé.")
        return redirect("scenarios_page")
    except ProtectedError:
        # Most common case: Scenario is linked to Backtest(s) via on_delete=PROTECT.
        linked_bts = list(
            Backtest.objects.filter(scenario=scenario).only("id", "name", "status").order_by("-created_at")[:50]
        )
        if linked_bts:
            lines = ", ".join([f"#{bt.id} {bt.name} ({bt.status})" for bt in linked_bts])
            more = "" if Backtest.objects.filter(scenario=scenario).count() <= 50 else " …"
            messages.error(
                request,
                "Impossible de supprimer ce scénario car il est lié à des backtests. "
                "Supprime d'abord les backtests associés (ou change leur scénario), puis réessaie. "
                f"Backtests liés: {lines}{more}",
            )
        else:
            messages.error(
                request,
                "Impossible de supprimer ce scénario car des objets liés existent (suppression protégée).",
            )
        return redirect("scenarios_page")


@login_required
@require_POST
def backtest_fetch_data(request, pk: int):
    """Launch a scoped market data fetch for a given backtest universe."""
    bt = get_object_or_404(Backtest, pk=pk)

    # Keep the universe snapshot aligned with the current Scenario definition
    # whenever the user triggers a new processing action.
    if bt.status != Backtest.Status.RUNNING:
        _refresh_backtest_universe_snapshot(bt)

    # Determine universe tickers from snapshot (preferred) or scenario
    tickers = []
    try:
        for r in (bt.universe_snapshot or []):
            if isinstance(r, dict):
                t = r.get("ticker")
                if t:
                    tickers.append(t)
    except Exception:
        tickers = []

    symbol_ids = list(Symbol.objects.filter(ticker__in=tickers).values_list("id", flat=True)) if tickers else list(bt.scenario.symbols.values_list("id", flat=True))

    from core.tasks import fetch_daily_bars_job_task

    # Create a PENDING job immediately so the UI can show 'en attente' even before the worker starts
    pj = ProcessingJob.objects.create(
        job_type=ProcessingJob.JobType.FETCH_BARS,
        status=ProcessingJob.Status.PENDING,
        backtest=bt,
        scenario=bt.scenario,
        created_by=request.user if request.user.is_authenticated else None,
        message="En attente d'exécution",
    )
    async_res = fetch_daily_bars_job_task.delay(
        symbol_ids=symbol_ids,
        scenario_id=bt.scenario_id,
        backtest_id=bt.id,
        user_id=request.user.id if request.user.is_authenticated else None,
        job_id=pj.id,
    )
    pj.task_id = getattr(async_res, 'id', '') or ''
    pj.save(update_fields=['task_id'])

    messages.success(request, "Collecte des données marché demandée pour ce backtest (traitement en arrière-plan).")
    return redirect("backtest_detail", pk=pk)


@login_required
@require_POST
def backtest_compute_metrics(request, pk: int):
    """Launch a scoped metrics computation for the backtest scenario + universe."""
    bt = get_object_or_404(Backtest, pk=pk)

    # Align snapshot with scenario universe before running compute.
    if bt.status != Backtest.Status.RUNNING:
        _refresh_backtest_universe_snapshot(bt)

    if bt.status != Backtest.Status.RUNNING:
        _refresh_backtest_universe_snapshot(bt)
    tickers = []
    try:
        for r in (bt.universe_snapshot or []):
            if isinstance(r, dict):
                t = r.get("ticker")
                if t:
                    tickers.append(t)
    except Exception:
        tickers = []

    symbol_ids = list(Symbol.objects.filter(ticker__in=tickers).values_list("id", flat=True)) if tickers else list(bt.scenario.symbols.values_list("id", flat=True))

    from core.tasks import compute_metrics_job_task

    pj = ProcessingJob.objects.create(
        job_type=ProcessingJob.JobType.COMPUTE_METRICS,
        status=ProcessingJob.Status.PENDING,
        backtest=bt,
        scenario=bt.scenario,
        created_by=request.user if request.user.is_authenticated else None,
        message="En attente d'exécution",
    )
    async_res = compute_metrics_job_task.delay(
        scenario_id=bt.scenario_id,
        symbol_ids=symbol_ids,
        recompute_all=False,
        backtest_id=bt.id,
        user_id=request.user.id if request.user.is_authenticated else None,
        job_id=pj.id,
    )
    pj.task_id = getattr(async_res, 'id', '') or ''
    pj.save(update_fields=['task_id'])

    messages.success(request, "Calcul des indicateurs demandé pour ce backtest (traitement en arrière-plan).")
    return redirect("backtest_detail", pk=pk)



@login_required
def email_settings_page(request):
    recipients = EmailRecipient.objects.all().order_by("-active", "email")
    settings_obj = EmailSettings.get_solo()
    scenarios = Scenario.objects.filter(active=True).order_by("name")

    if request.method == "POST" and request.POST.get("_action") == "add_recipient":
        form = EmailRecipientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Destinataire ajouté.")
            return redirect("email_settings")
        messages.error(request, "Email invalide.")
        settings_form = EmailSettingsForm(instance=settings_obj)

    elif request.method == "POST" and request.POST.get("_action") == "save_settings":
        settings_form = EmailSettingsForm(request.POST, instance=settings_obj)
        if settings_form.is_valid():
            settings_form.save()
            messages.success(request, "Paramètres email mis à jour.")
            return redirect("email_settings")
        messages.error(request, "Paramètres invalides.")
        form = EmailRecipientForm()

    else:
        form = EmailRecipientForm()
        settings_form = EmailSettingsForm(instance=settings_obj)

    return render(
        request,
        "email_settings.html",
        {
            "recipients": recipients,
            "form": form,
            "settings_form": settings_form,
            "settings_obj": settings_obj,
            "scenarios": scenarios,
        },
    )


@login_required
@require_POST
@login_required
@require_POST
def run_compute_now(request):
    """Run an incremental compute.

    UI can pass scenario_id to scope compute to a single scenario.
    If scenario_id == 'ALL', keep legacy behavior (all scenarios).
    """
    scenario_id = (request.POST.get("scenario_id") or "").strip()
    try:
        if not scenario_id:
            messages.error(request, "Veuillez choisir un scénario (ou 'Tous les scénarios').")
            return redirect("email_settings")

        if scenario_id.upper() == "ALL":
            from core.tasks import compute_metrics_all_job_task
            compute_metrics_all_job_task.delay(False, user_id=request.user.id if request.user.is_authenticated else None)
            messages.success(request, "Calculs demandés (tous scénarios, traitement en arrière-plan).")
        else:
            from core.tasks import compute_metrics_job_task
            compute_metrics_job_task.delay(
                scenario_id=int(scenario_id),
                symbol_ids=None,
                recompute_all=False,
                backtest_id=None,
                user_id=request.user.id if request.user.is_authenticated else None,
                job_id=None,
            )
            messages.success(request, "Calculs demandés (scénario sélectionné, traitement en arrière-plan).")
    except Exception as e:
        messages.error(request, f"Erreur lancement calculs: {e}")
    return redirect("email_settings")


@login_required
@require_POST
def run_recompute_all_now(request):
    """Force a full recompute.

    UI can pass scenario_id to scope recompute to a single scenario.
    If scenario_id == 'ALL' (or missing), keep legacy behavior (all scenarios).
    """
    scenario_id = (request.POST.get("scenario_id") or "").strip()
    try:
        if not scenario_id:
            messages.error(request, "Veuillez choisir un scénario (ou 'Tous les scénarios').")
            return redirect("email_settings")

        if scenario_id.upper() == "ALL":
            from core.tasks import compute_metrics_all_job_task
            compute_metrics_all_job_task.delay(True, user_id=request.user.id if request.user.is_authenticated else None)
            messages.success(request, "Recompute complet demandé (tous scénarios).")
        else:
            from core.tasks import compute_metrics_job_task
            compute_metrics_job_task.delay(
                scenario_id=int(scenario_id),
                symbol_ids=None,
                recompute_all=True,
                backtest_id=None,
                user_id=request.user.id if request.user.is_authenticated else None,
                job_id=None,
            )
            messages.success(request, "Recompute complet demandé (scénario sélectionné).")
    except Exception as e:
        messages.error(request, f"Erreur recompute complet: {e}")
    return redirect("email_settings")


@login_required
@require_POST
def backtest_recompute_metrics(request, pk: int):
    """Launch a full recompute scoped to the backtest scenario + universe."""
    bt = get_object_or_404(Backtest, pk=pk)

    # Keep reproducibility: align snapshot before running.
    if bt.status != Backtest.Status.RUNNING:
        _refresh_backtest_universe_snapshot(bt)

    tickers = []
    try:
        for r in (bt.universe_snapshot or []):
            if isinstance(r, dict):
                t = r.get("ticker")
                if t:
                    tickers.append(t)
    except Exception:
        tickers = []

    symbol_ids = (
        list(Symbol.objects.filter(ticker__in=tickers).values_list("id", flat=True))
        if tickers
        else list(bt.scenario.symbols.values_list("id", flat=True))
    )

    from core.tasks import compute_metrics_job_task

    pj = ProcessingJob.objects.create(
        job_type=ProcessingJob.JobType.COMPUTE_METRICS,
        status=ProcessingJob.Status.PENDING,
        backtest=bt,
        scenario=bt.scenario,
        created_by=request.user if request.user.is_authenticated else None,
        message="En attente d'exécution (recompute complet scénario)",
    )
    async_res = compute_metrics_job_task.delay(
        scenario_id=bt.scenario_id,
        symbol_ids=symbol_ids,
        recompute_all=True,
        backtest_id=bt.id,
        user_id=request.user.id if request.user.is_authenticated else None,
        job_id=pj.id,
    )
    pj.task_id = getattr(async_res, "id", "") or ""
    pj.save(update_fields=["task_id"])

    messages.success(request, "Recompute complet demandé pour ce backtest (scénario uniquement).")
    return redirect("backtest_detail", pk=pk)


@login_required
@require_POST
def fetch_bars_now(request):
    """Fetch daily bars immediately (useful for manual refresh)."""
    try:
        from core.tasks import fetch_daily_bars_job_task
        fetch_daily_bars_job_task.delay(user_id=request.user.id if request.user.is_authenticated else None)
        JobLog.objects.create(level="INFO", job="fetch_bars", message="Fetch daily bars demandé (Celery + job tracking).")
        messages.success(request, "Collecte demandée (traitement en arrière-plan).")
    except Exception as e:
        JobLog.objects.create(level="ERROR", job="fetch_bars", message="Erreur lancement fetch", traceback=str(e))
        messages.error(request, f"Erreur lancement collecte: {e}")
    return redirect("email_settings")


@login_required
@require_POST
def send_mail_now(request):
    try:
        from core.tasks import send_daily_alerts_task
        send_daily_alerts_task.delay()
        messages.success(request, "Envoi email demandé (en background via Celery).")
    except Exception as e:
        messages.error(request, f"Erreur envoi email: {e}")
    return redirect("email_settings")

@login_required
@require_POST
def email_recipient_toggle(request, pk: int):
    r = get_object_or_404(EmailRecipient, pk=pk)
    r.active = not r.active
    r.save(update_fields=["active"])
    return redirect("email_settings")


@login_required
@require_POST
def email_recipient_delete(request, pk: int):
    r = get_object_or_404(EmailRecipient, pk=pk)
    r.delete()
    return redirect("email_settings")


@login_required
def api_symbol_search(request):
    q = (request.GET.get("q") or "").strip()
    if len(q) < 1:
        return JsonResponse({"data": []})
    client = TwelveDataClient()
    try:
        items = client.symbol_search(q, limit=12)
    except Exception as e:
        return JsonResponse({"error": str(e), "data": []}, status=400)
    out = []
    for it in items:
        out.append(
            {
                "symbol": it.get("symbol") or it.get("ticker") or "",
                "exchange": it.get("exchange") or "",
                "name": it.get("instrument_name") or it.get("name") or "",
                "instrument_type": it.get("instrument_type") or "",
                "country": it.get("country") or "",
                "currency": it.get("currency") or "",
            }
        )
    return JsonResponse({"data": out})

from django.core.paginator import Paginator


@login_required
def logs_page(request):
    """Log viewer to help debug async/background jobs."""
    level = (request.GET.get("level") or "").upper().strip()
    job = (request.GET.get("job") or "").strip()

    qs = JobLog.objects.all().order_by("-created_at")
    if level in {"INFO", "WARNING", "ERROR"}:
        qs = qs.filter(level=level)
    if job:
        qs = qs.filter(job__icontains=job)

    paginator = Paginator(qs, 50)
    page = paginator.get_page(request.GET.get("page"))
    return render(request, "logs.html", {"page": page, "level": level, "job": job})


@login_required
def jobs_page(request):
    """List background processing jobs (pending/running/done/failed)."""
    status = (request.GET.get("status") or "").strip().upper()
    job_type = (request.GET.get("type") or "").strip().upper()

    qs = ProcessingJob.objects.select_related("backtest", "scenario", "created_by")
    if status in {"PENDING", "RUNNING", "DONE", "FAILED"}:
        qs = qs.filter(status=status)
    if job_type:
        qs = qs.filter(job_type=job_type)

    # Counters for quick navigation
    counts = {
        "PENDING": ProcessingJob.objects.filter(status=ProcessingJob.Status.PENDING).count(),
        "RUNNING": ProcessingJob.objects.filter(status=ProcessingJob.Status.RUNNING).count(),
        "DONE": ProcessingJob.objects.filter(status=ProcessingJob.Status.DONE).count(),
        "FAILED": ProcessingJob.objects.filter(status=ProcessingJob.Status.FAILED).count(),
    }

    jobs = qs.order_by("-created_at")[:200]
    return render(
        request,
        "jobs.html",
        {
            "jobs": jobs,
            "status": status,
            "job_type": job_type,
            "counts": counts,
            "job_types": ProcessingJob.JobType.choices,
        },
    )



@login_required
def backtests_page(request):
    """Archive page: list saved backtests with a simple search."""
    qs = Backtest.objects.select_related("scenario").all()
    q = (request.GET.get("q") or "").strip()
    scenario_id = (request.GET.get("scenario") or "").strip()

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q))
    if scenario_id:
        qs = qs.filter(scenario_id=scenario_id)

    scenarios = Scenario.objects.all().order_by("name")
    return render(
        request,
        "backtests_list.html",
        {"backtests": qs[:200], "q": q, "scenarios": scenarios, "scenario_id": scenario_id},
    )


@login_required
def backtest_create(request):
    """Create a Backtest configuration (no computation in Feature 1)."""
    if request.method == "POST":
        form = BacktestForm(request.POST)
        if form.is_valid():
            bt = form.save(commit=False)
            bt.created_by = request.user if request.user.is_authenticated else None
            bt.save()
            _refresh_backtest_universe_snapshot(bt)
            messages.success(request, "Backtest enregistré (configuration).")
            return redirect("backtest_detail", pk=bt.pk)
    else:
        form = BacktestForm(initial={"include_all_tickers": True})
    signal_lines_json = json.dumps(form["signal_lines"].value() or [])
    return render(request, "backtest_create.html", {"form": form, "signal_choices_json": json.dumps(BACKTEST_SIGNAL_CHOICES), "signal_lines_json": signal_lines_json})


@login_required
def backtest_update(request, pk: int):
    """Edit an existing Backtest configuration.

    Important:
    - Editing parameters can invalidate existing results.
    - We therefore reset stored results + portfolio tables on save.
    - Universe snapshot is refreshed from the (possibly new) scenario.
    """
    bt = get_object_or_404(Backtest.objects.select_related("scenario"), pk=pk)

    if bt.status == Backtest.Status.RUNNING:
        messages.error(request, "Impossible de modifier un backtest en cours d'exécution. Attends la fin du traitement.")
        return redirect("backtest_detail", pk=pk)

    if request.method == "POST":
        form = BacktestForm(request.POST, instance=bt)
        if form.is_valid():
            bt = form.save(commit=False)
            # Reset computed results (they are now stale)
            bt.results = {}
            bt.status = Backtest.Status.PENDING
            bt.error_message = ""
            bt.save()

            # Clear persisted portfolio aggregates (they will be recomputed at next run)
            try:
                bt.portfolio_daily.all().delete()
            except Exception:
                pass
            try:
                if hasattr(bt, "portfolio_kpi") and bt.portfolio_kpi:
                    bt.portfolio_kpi.delete()
            except Exception:
                pass

            # Refresh universe snapshot (scenario may have changed)
            _refresh_backtest_universe_snapshot(bt)

            messages.success(request, "Backtest mis à jour. Les résultats précédents ont été réinitialisés (relance Fetch/Compute/Run).")
            return redirect("backtest_detail", pk=pk)
    else:
        form = BacktestForm(instance=bt)

    signal_lines_json = json.dumps(form["signal_lines"].value() or [])

    return render(
        request,
        "backtest_edit.html",
        {
            "form": form,
            "bt": bt,
            "signal_choices_json": json.dumps(BACKTEST_SIGNAL_CHOICES),
            "signal_lines_json": signal_lines_json,
        },
    )


@login_required
def backtest_detail(request, pk: int):
    bt = get_object_or_404(Backtest.objects.select_related("scenario"), pk=pk)
    jobs = ProcessingJob.objects.filter(backtest=bt).order_by("-created_at")[:30]

    latest_by_type = {}
    for jt in [ProcessingJob.JobType.FETCH_BARS, ProcessingJob.JobType.COMPUTE_METRICS, ProcessingJob.JobType.RUN_BACKTEST]:
        latest_by_type[jt] = ProcessingJob.objects.filter(backtest=bt, job_type=jt).order_by("-created_at").first()

    return render(request, "backtest_detail.html", {"bt": bt, "jobs": jobs, "latest_by_type": latest_by_type})


@login_required
@require_POST
def backtest_delete(request, pk: int):
    """Delete a backtest and its persisted results.

    Notes:
    - Related portfolio daily/KPI are CASCADE -> removed automatically.
    - We prevent deletion while RUNNING to avoid celery tasks writing to a deleted row.
    """
    bt = get_object_or_404(Backtest, pk=pk)
    if bt.status == Backtest.Status.RUNNING:
        messages.error(request, "Impossible de supprimer un backtest en cours d'exécution. Attends la fin ou relance plus tard.")
        return redirect("backtest_detail", pk=pk)

    name = bt.name
    bt.delete()
    messages.success(request, f"Backtest supprimé : {name}.")
    return redirect("backtests_page")


@login_required
def backtest_run(request, pk: int):
    """Launch a backtest run asynchronously."""
    bt = get_object_or_404(Backtest, pk=pk)
    if request.method != "POST":
        return redirect("backtest_detail", pk=pk)

    # Refresh universe snapshot from the current scenario before running.
    # This ensures that scenario changes are taken into account when re-running a backtest.
    if bt.status != Backtest.Status.RUNNING:
        _refresh_backtest_universe_snapshot(bt)

    Backtest.objects.filter(id=bt.id).update(status=Backtest.Status.PENDING, error_message="")
    from .tasks import run_backtest_job_task

    pj = ProcessingJob.objects.create(
        job_type=ProcessingJob.JobType.RUN_BACKTEST,
        status=ProcessingJob.Status.PENDING,
        backtest=bt,
        scenario=bt.scenario,
        created_by=request.user if request.user.is_authenticated else None,
        message="En attente d'exécution",
    )
    async_res = run_backtest_job_task.delay(bt.id, user_id=request.user.id if request.user.is_authenticated else None, job_id=pj.id)
    pj.task_id = getattr(async_res, 'id', '') or ''
    pj.save(update_fields=['task_id'])
    messages.success(request, "Backtest lancé (traitement en arrière-plan).")
    return redirect("backtest_detail", pk=pk)

@login_required
def backtest_results(request, pk: int):
    """Readable results view for a computed backtest.

    Uses the JSON stored in Backtest.results (no recomputation here).
    """
    bt = get_object_or_404(Backtest.objects.select_related("scenario"), pk=pk)
    results = bt.results or {}
    tickers_map = results.get("tickers") or {}

    if not tickers_map:
        messages.warning(request, "Aucun résultat disponible pour ce backtest (lance-le d'abord).")
        return redirect("backtest_detail", pk=pk)

    # Selected ticker / line
    ticker = request.GET.get("ticker") or next(iter(tickers_map.keys()))
    tentry = tickers_map.get(ticker) or next(iter(tickers_map.values()))
    ticker = ticker if ticker in tickers_map else next(iter(tickers_map.keys()))

    try:
        line_index = int(request.GET.get("line", "1"))
    except ValueError:
        line_index = 1

    lines = tentry.get("lines") or []
    line = next((l for l in lines if int(l.get("line_index", 0)) == line_index), None)
    if line is None and lines:
        line = lines[0]
        line_index = int(line.get("line_index", 1))

    daily = (line or {}).get("daily") or []
    final = (line or {}).get("final") or {}

    # Portfolio synthesis (Feature 8)
    portfolio = results.get("portfolio") or {}
    port_kpi = portfolio.get("kpi") or {}
    port_daily = portfolio.get("daily") or []
    # keep charts responsive: last 400 points is enough for UI
    if len(port_daily) > 400:
        port_daily_for_ui = port_daily[-400:]
    else:
        port_daily_for_ui = port_daily

    # Truncate very large series for UI rendering (default: last 200 days)
    show_all = request.GET.get("all") == "1"
    limit = 200
    total_daily_count = len(daily)
    is_truncated = (total_daily_count > limit) and (not show_all)
    if is_truncated:
        daily = daily[-limit:]


    # For dropdowns in UI
    ticker_options = []
    for tk, te in tickers_map.items():
        for l in (te.get("lines") or []):
            ticker_options.append({
                "ticker": tk,
                "line_index": l.get("line_index"),
                "buy": l.get("buy"),
                "sell": l.get("sell"),
            })

    return render(
        request,
        "backtest_results.html",
        {
            "bt": bt,
            "results": results,
            "ticker": ticker,
            "line_index": line_index,
            "line": line,
            "daily": daily,
            "daily_json": json.dumps(daily),
            "final": final,
            "portfolio_kpi": port_kpi,
            "portfolio_daily": port_daily_for_ui,
            "portfolio_daily_json": json.dumps(port_daily_for_ui),
            "is_truncated": is_truncated,
            "total_daily_count": total_daily_count,
            "ticker_options": ticker_options,
            "portfolio_kpi": port_kpi,
            "portfolio_daily": port_daily_for_ui,
            "portfolio_daily_json": json.dumps(port_daily_for_ui),
        },
    )


@login_required
def backtest_export_debug_csv(request, pk: int):
    """Export a debug CSV for one (ticker, line) from Backtest.results."""
    bt = get_object_or_404(Backtest, pk=pk)
    results = bt.results or {}
    tickers_map = results.get("tickers") or {}
    if not tickers_map:
        messages.warning(request, "Aucun résultat à exporter (lance le backtest).")
        return redirect("backtest_detail", pk=pk)

    ticker = request.GET.get("ticker") or next(iter(tickers_map.keys()))
    tentry = tickers_map.get(ticker) or next(iter(tickers_map.values()))
    ticker = ticker if ticker in tickers_map else next(iter(tickers_map.keys()))

    try:
        line_index = int(request.GET.get("line", "1"))
    except ValueError:
        line_index = 1
    lines = tentry.get("lines") or []
    line = next((l for l in lines if int(l.get("line_index", 0)) == line_index), None)
    if line is None and lines:
        line = lines[0]
        line_index = int(line.get("line_index", 1))

    daily = (line or {}).get("daily") or []
    if not daily:
        messages.warning(request, "Pas de données journalières pour cet export.")
        return redirect("backtest_results", pk=pk)

    # Build a stable header (union of keys)
    header_keys = []
    seen = set()
    for row in daily:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                header_keys.append(k)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"backtest_{bt.id}_{ticker}_L{line_index}_debug.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(header_keys)
    for row in daily:
        writer.writerow([row.get(k, "") for k in header_keys])

    return response


@login_required
def backtest_export_excel(request, pk: int):
    """Export full backtest results to Excel (settings + universe + summary + daily sheets + charts)."""
    bt = get_object_or_404(Backtest, pk=pk)
    results = bt.results or {}
    tickers_map = results.get("tickers") or {}
    if not tickers_map:
        messages.warning(request, "Aucun résultat à exporter (lance le backtest).")
        return redirect("backtest_detail", pk=pk)

    # --- Volume guards (optional) ---
    num_tickers_total = len(tickers_map)
    limit_excel_details = should_limit_excel(num_tickers_total)
    top_n = excel_top_n()
    selected_tickers_for_details = list(tickers_map.keys())
    if limit_excel_details:
        selected_tickers_for_details = select_top_tickers_by_metric(tickers_map, top_n)

    def _to_float(x):
        if x is None or x == "":
            return None
        try:
            return float(x)
        except Exception:
            try:
                return float(str(x))
            except Exception:
                return None

    def _pct_ratio_to_percent(x):
        """Stored as ratio (0.01==1%) -> percent value (1.0)."""
        f = _to_float(x)
        return None if f is None else f * 100.0

    def _auto_width(ws, max_col=40):
        for col in range(1, min(ws.max_column, max_col) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 55)

    wb = Workbook()
    wb.remove(wb.active)

    # --- Settings ---
    ws = wb.create_sheet("Settings")
    ws.append(["Clé", "Valeur"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    meta = results.get("meta") or {}
    settings_rows = [
        ("Backtest ID", bt.id),
        ("Nom", bt.name),
        ("Description", bt.description or ""),
        ("Scénario", getattr(bt.scenario, "name", "") if bt.scenario_id else ""),
        ("Période début", bt.start_date.isoformat() if bt.start_date else ""),
        ("Période fin", bt.end_date.isoformat() if bt.end_date else ""),
        ("CP (capital total)", bt.capital_total),
        ("CT (capital par ticker)", bt.capital_per_ticker),
        ("X (seuil ratio_p %)", bt.ratio_threshold),
        ("Clôture fin backtest", "Oui" if bt.close_positions_at_end else "Non"),
        ("Statut", bt.status),
        ("global_cash_end", meta.get("global_cash_end", "")),
        ("engine_version", meta.get("engine_version", "")),
    ]
    for k, v in settings_rows:
        ws.append([k, v])
    _auto_width(ws)

    # --- Universe (snapshot) ---
    ws_u = wb.create_sheet("Universe")
    ws_u.append(["Ticker", "Exchange"])
    ws_u["A1"].font = Font(bold=True)
    ws_u["B1"].font = Font(bold=True)

    uni = bt.universe_snapshot or []
    if isinstance(uni, list):
        for item in uni:
            if isinstance(item, dict):
                ws_u.append([item.get("ticker", ""), item.get("exchange", "")])
            else:
                ws_u.append([str(item), ""]) 
    _auto_width(ws_u)

    # --- Summary ---
    ws_s = wb.create_sheet("Summary")
    ws_s.append([
        "Ticker",
        "Line #",
        "BUY",
        "SELL",
        "Allocated",
        "N",
        "S_G_N (%)",
        "BT (%)",
        "NB_JOUR_OUVRES",
        "BMJ (%)",
        "BMD (%)",
        "BUY_DAYS_CLOSED",
        "Cash end",
    ])
    ws_s.freeze_panes = "A2"
    for cell in ws_s[1]:
        cell.font = Font(bold=True)

    for ticker in selected_tickers_for_details:
        tentry = tickers_map.get(ticker) or {}
        for line in (tentry or {}).get("lines") or []:
            fin = line.get("final") or {}
            ws_s.append([
                ticker,
                line.get("line_index"),
                line.get("buy"),
                line.get("sell"),
                "Oui" if line.get("allocated") else "Non",
                fin.get("N"),
                _pct_ratio_to_percent(fin.get("S_G_N")),
                _pct_ratio_to_percent(fin.get("BT")),
                fin.get("NB_JOUR_OUVRES"),
                _pct_ratio_to_percent(fin.get("BMJ")),
                _pct_ratio_to_percent(fin.get("BMD")),
                fin.get("BUY_DAYS_CLOSED"),
                _to_float(fin.get("cash_ticker_end")),
            ])
    _auto_width(ws_s)

    # --- Portfolio (Feature 8) ---
    portfolio = results.get("portfolio") or {}
    port_kpi = portfolio.get("kpi") or {}
    port_daily = portfolio.get("daily") or []

    ws_p = wb.create_sheet("Portfolio")
    ws_p.append(["Clé", "Valeur"])
    ws_p["A1"].font = Font(bold=True)
    ws_p["B1"].font = Font(bold=True)
    for k, v in [
        ("capital_total", port_kpi.get("capital_total")),
        ("invested_end", port_kpi.get("invested_end")),
        ("equity_end", port_kpi.get("equity_end")),
        ("BT_return", _pct_ratio_to_percent(port_kpi.get("BT"))),
        ("BMJ_return", _pct_ratio_to_percent(port_kpi.get("BMJ"))),
        ("NB_days", port_kpi.get("NB_DAYS")),
        ("max_drawdown", _pct_ratio_to_percent(port_kpi.get("max_drawdown"))),
    ]:
        ws_p.append([k, v])

    ws_pd = wb.create_sheet("Portfolio_Daily")
    ws_pd.append(["Date", "Equity", "Invested", "GlobalCash", "CashAllocated", "PositionsValue", "Drawdown (%)"])
    ws_pd.freeze_panes = "A2"
    for cell in ws_pd[1]:
        cell.font = Font(bold=True)
    for r in port_daily:
        ws_pd.append([
            r.get("date"),
            _to_float(r.get("equity")),
            _to_float(r.get("invested")),
            _to_float(r.get("global_cash")),
            _to_float(r.get("cash_allocated")),
            _to_float(r.get("positions_value")),
            _pct_ratio_to_percent(r.get("drawdown")),
        ])

    # Equity chart
    try:
        from openpyxl.chart import LineChart, Reference

        if len(port_daily) >= 2:
            chart = LineChart()
            chart.title = "Equity curve"
            chart.y_axis.title = "Equity"
            chart.x_axis.title = "Date"
            data_ref = Reference(ws_pd, min_col=2, min_row=1, max_col=2, max_row=1 + len(port_daily))
            cats_ref = Reference(ws_pd, min_col=1, min_row=2, max_row=1 + len(port_daily))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 12
            chart.width = 28
            ws_pd.add_chart(chart, "I2")
    except Exception:
        pass

    _auto_width(ws_p)
    _auto_width(ws_pd)

    # --- Daily sheets + charts ---
    # If volume guards are enabled and the universe is too large, we only include Top N tickers' daily sheets.
    tickers_for_daily = selected_tickers_for_details
    for ticker in tickers_for_daily:
        tentry = tickers_map.get(ticker) or {}
        for line in (tentry or {}).get("lines") or []:
            li = int(line.get("line_index") or 1)
            ws_name = f"{ticker}_L{li}"[:31]
            ws_d = wb.create_sheet(ws_name)

            ws_d.append([
                "Date",
                "Close",
                "Prix_vert",
                "Prix_rouge",
                "Ratio_p (%)",
                "Tradable",
                "Alerts",
                "Action",
                "G (%)",
                "N",
                "S_G_N (%)",
                "BT (%)",
                "NB_JOUR_OUVRES",
                "BMJ (%)",
                "BMD (%)",
                "BUY_DAYS_CLOSED",
                "Cash",
                "Shares",
            ])
            ws_d.freeze_panes = "A2"
            for cell in ws_d[1]:
                cell.font = Font(bold=True)

            daily = line.get("daily") or []
            for r in daily:
                close_px = _to_float(r.get("price_close"))
                shares = _to_float(r.get("shares")) or 0
                in_pos = shares > 0
                ws_d.append([
                    r.get("date"),
                    close_px,
                    close_px if in_pos else None,
                    close_px if (not in_pos) else None,
                    _to_float(r.get("ratio_P_pct")),  # already 0-100
                    "Oui" if r.get("tradable") else "Non",
                    ",".join(r.get("alerts") or []),
                    r.get("action") or "",
                    _pct_ratio_to_percent(r.get("action_G")),
                    r.get("N"),
                    _pct_ratio_to_percent(r.get("S_G_N")),
                    _pct_ratio_to_percent(r.get("BT")),
                    r.get("NB_JOUR_OUVRES"),
                    _pct_ratio_to_percent(r.get("BMJ")),
                    _pct_ratio_to_percent(r.get("BMD")),
                    r.get("BUY_DAYS_CLOSED"),
                    _to_float(r.get("cash_ticker")),
                    r.get("shares"),
                ])

            _auto_width(ws_d, max_col=20)

            if ws_d.max_row >= 3:
                chart = LineChart()
                chart.title = f"{ticker} L{li} - S_G_N / BT / BMJ / BMD (%)"
                chart.y_axis.title = "%"
                chart.x_axis.title = "Date"

                # Data columns (after adding Prix_vert/Prix_rouge):
                # S_G_N=11, BT=12, BMJ=14, BMD=15
                for col in (11, 12, 14, 15):
                    data = Reference(ws_d, min_col=col, min_row=1, max_row=ws_d.max_row)
                    chart.add_data(data, titles_from_data=True)

                cats = Reference(ws_d, min_col=1, min_row=2, max_row=ws_d.max_row)
                chart.set_categories(cats)
                chart.height = 12
                chart.width = 28
                ws_d.add_chart(chart, f"A{ws_d.max_row + 3}")

                # Price + position chart (stacked area):
                # - Prix_vert populated when shares > 0
                # - Prix_rouge populated when shares == 0
                # The user expects an "aire empilée" chart in Excel.
                try:
                    from openpyxl.chart import AreaChart
                    from openpyxl.chart.series import DataPoint
                    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

                    chart2 = AreaChart()
                    chart2.grouping = "stacked"
                    chart2.title = f"{ticker} L{li} - Price + Position (aire empilée)"
                    chart2.y_axis.title = "Price"
                    chart2.x_axis.title = "Date"

                    # Data columns: Prix_vert=3, Prix_rouge=4 (since we inserted 2 cols after Close)
                    data2 = Reference(ws_d, min_col=3, min_row=1, max_col=4, max_row=ws_d.max_row)
                    chart2.add_data(data2, titles_from_data=True)
                    chart2.set_categories(cats)
                    chart2.height = 12
                    chart2.width = 28

                    # Set fills to match semantic: green=in position, red=out
                    # (If Excel theme overrides, it still remains readable thanks to legend.)
                    try:
                        if chart2.series and len(chart2.series) >= 2:
                            chart2.series[0].graphicalProperties.solidFill = "00B400"  # green
                            chart2.series[0].graphicalProperties.line.solidFill = "00B400"
                            chart2.series[1].graphicalProperties.solidFill = "DC0000"  # red
                            chart2.series[1].graphicalProperties.line.solidFill = "DC0000"
                    except Exception:
                        pass

                    ws_d.add_chart(chart2, f"I{ws_d.max_row + 3}")
                except Exception:
                    pass

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="backtest_{bt.id}_export.xlsx"'
    return response


@login_required
def backtest_export_excel_compact(request, pk: int):
    """Google Sheets-friendly compact export.

    This export is designed for many tickers:
    - Settings / Universe / Summary sheets
    - A single "Daily" sheet containing all daily rows across (ticker, line)
    - Charts are embedded as PNG images (Sheets displays images reliably)

    Query params:
      - charts=0|1  (default 1) include charts as images
      - chart_mode=top|all|first
          * top   : (default) génère des charts pour les meilleurs couples (ticker, ligne) selon BT final
          * all   : génère des charts pour tous les couples (ticker, ligne), avec limite de sécurité
          * first : un seul couple (ticker, ligne) (comportement historique)
      - chart_limit=6 (optionnel, défaut 6) nombre max de graphiques à insérer dans l'onglet Charts
      - chart_ticker=MSFT (optionnel, utilisé uniquement si chart_mode=first)
      - chart_line=1 (optionnel, utilisé uniquement si chart_mode=first)
    """

    bt = get_object_or_404(Backtest, pk=pk)
    results = bt.results or {}
    tickers_map = results.get("tickers") or {}
    if not tickers_map:
        messages.warning(request, "Aucun résultat à exporter (lance le backtest).")
        return redirect("backtest_detail", pk=pk)

    # --- Volume guards (optional) ---
    num_tickers_total = len(tickers_map)
    limit_compact_daily = should_limit_excel(num_tickers_total)
    top_n = excel_top_n()
    selected_tickers_for_details = list(tickers_map.keys())
    if limit_compact_daily:
        selected_tickers_for_details = select_top_tickers_by_metric(tickers_map, top_n)

    charts_enabled = request.GET.get("charts", "1") != "0"
    chart_mode = (request.GET.get("chart_mode") or "top").lower().strip()
    try:
        chart_limit = int(request.GET.get("chart_limit", "6"))
    except Exception:
        chart_limit = 6

    chart_ticker = request.GET.get("chart_ticker")
    try:
        chart_line = int(request.GET.get("chart_line", "1"))
    except Exception:
        chart_line = 1

    def _to_float(x):
        if x is None or x == "":
            return None
        try:
            return float(x)
        except Exception:
            try:
                return float(str(x))
            except Exception:
                return None

    def _pct(x):
        """Stored as ratio (e.g. 0.0123) -> percent (1.23)."""
        f = _to_float(x)
        return None if f is None else f * 100.0

    def _auto_width(ws, max_col=30):
        for col in range(1, min(ws.max_column, max_col) + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 45)

    wb = Workbook()
    wb.remove(wb.active)

    # -------- Settings --------
    ws = wb.create_sheet("Settings")
    ws.append(["Clé", "Valeur"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    meta = results.get("meta") or {}
    rows = [
        ("Backtest ID", bt.id),
        ("Nom", bt.name),
        ("Description", bt.description or ""),
        ("Scénario", getattr(bt.scenario, "name", "") if bt.scenario_id else ""),
        ("Période début", bt.start_date.isoformat() if bt.start_date else ""),
        ("Période fin", bt.end_date.isoformat() if bt.end_date else ""),
        ("CP (capital total)", bt.capital_total),
        ("CT (capital par ticker)", bt.capital_per_ticker),
        ("X (seuil ratio_p %)", bt.ratio_threshold),
        ("Clôture fin backtest", "Oui" if bt.close_positions_at_end else "Non"),
        ("Statut", bt.status),
        ("global_cash_end", meta.get("global_cash_end", "")),
        ("engine_version", meta.get("engine_version", "")),
    ]
    for k, v in rows:
        ws.append([k, v])
    _auto_width(ws)

    # -------- Universe --------
    ws_u = wb.create_sheet("Universe")
    ws_u.append(["Ticker", "Exchange"])
    ws_u["A1"].font = Font(bold=True)
    ws_u["B1"].font = Font(bold=True)
    uni = bt.universe_snapshot or []
    if isinstance(uni, list):
        for item in uni:
            if isinstance(item, dict):
                ws_u.append([item.get("ticker", ""), item.get("exchange", "")])
            else:
                ws_u.append([str(item), ""])
    _auto_width(ws_u)

    # -------- Summary --------
    ws_s = wb.create_sheet("Summary")
    ws_s.append([
        "Ticker", "Line #", "BUY", "SELL", "Allocated",
        "N", "S_G_N (%)", "BT (%)", "NB_JOUR_OUVRES", "BMJ (%)", "BMD (%)", "BUY_DAYS_CLOSED", "Cash end",
    ])
    ws_s.freeze_panes = "A2"
    for cell in ws_s[1]:
        cell.font = Font(bold=True)

    # Also build a flat list of all daily rows for compact sheet
    daily_rows = []

    # Pick a default chart target
    if not chart_ticker:
        chart_ticker = next(iter(tickers_map.keys()))

    for ticker, tentry in tickers_map.items():
        lines = (tentry or {}).get("lines") or []
        for strat in lines:
            sidx = int(strat.get("line_index", 0) or 0) or 1
            fin = strat.get("final") or {}
            ws_s.append([
                ticker,
                sidx,
                strat.get("buy"),
                strat.get("sell"),
                "Oui" if strat.get("allocated") else "Non",
                fin.get("N"),
                _pct(fin.get("S_G_N")),
                _pct(fin.get("BT")),
                fin.get("NB_JOUR_OUVRES"),
                _pct(fin.get("BMJ")),
                _to_float(fin.get("cash_ticker_end")),
            ])

            daily = strat.get("daily") or []
            for r in daily:
                daily_rows.append({
                    "date": r.get("date"),
                    "ticker": ticker,
                    "line": sidx,
                    "buy": strat.get("buy"),
                    "sell": strat.get("sell"),
                    "close": _to_float(r.get("price_close")),
                    "ratio_p_pct": _to_float(r.get("ratio_P_pct")),
                    "tradable": bool(r.get("tradable")),
                    "alerts": ",".join(r.get("alerts") or []),
                    "action": r.get("action") or "",
                    "G_pct": _pct(r.get("action_G")),
                    "N": r.get("N"),
                    "S_G_N_pct": _pct(r.get("S_G_N")),
                    "BT_pct": _pct(r.get("BT")),
                    "NB_JOUR_OUVRES": r.get("NB_JOUR_OUVRES"),
                    "BMJ_pct": _pct(r.get("BMJ")),
                    "cash": _to_float(r.get("cash_ticker")),
                    "shares": r.get("shares"),
                })

    _auto_width(ws_s)

    # -------- Daily (compact) --------
    ws_d = wb.create_sheet("Daily")
    header = [
        "Date", "Ticker", "Line #", "BUY", "SELL",
        "Close", "Ratio_p (%)", "Tradable", "Alerts", "Action",
        "G (%)", "N", "S_G_N (%)", "BT (%)",
        "NB_JOUR_OUVRES", "BMJ (%)", "BMD (%)", "BUY_DAYS_CLOSED", "Cash", "Shares",
    ]
    ws_d.append(header)
    ws_d.freeze_panes = "A2"
    for cell in ws_d[1]:
        cell.font = Font(bold=True)

    # stable ordering
    daily_rows.sort(key=lambda x: (x.get("date") or "", x.get("ticker") or "", x.get("line") or 0))
    for r in daily_rows:
        ws_d.append([
            r["date"], r["ticker"], r["line"], r["buy"], r["sell"],
            r["close"], r["ratio_p_pct"], "Oui" if r["tradable"] else "Non",
            r["alerts"], r["action"], r["G_pct"], r["N"], r["S_G_N_pct"], r["BT_pct"],
            r["NB_JOUR_OUVRES"], r["BMJ_pct"], r.get("BMD_pct"), r.get("BUY_DAYS_CLOSED"), r["cash"], r["shares"],
        ])
    _auto_width(ws_d, max_col=18)

    # -------- Portfolio (compact) – Feature 8 --------
    portfolio = results.get("portfolio") or {}
    port_kpi = portfolio.get("kpi") or {}
    port_daily = portfolio.get("daily") or []

    ws_p = wb.create_sheet("Portfolio")
    ws_p.append(["Clé", "Valeur"])
    ws_p["A1"].font = Font(bold=True)
    ws_p["B1"].font = Font(bold=True)
    for k, v in [
        ("capital_total", port_kpi.get("capital_total")),
        ("invested_end", port_kpi.get("invested_end")),
        ("equity_end", port_kpi.get("equity_end")),
        ("BT_return", _pct_ratio_to_percent(port_kpi.get("BT"))),
        ("BMJ_return", _pct_ratio_to_percent(port_kpi.get("BMJ"))),
        ("NB_days", port_kpi.get("NB_DAYS")),
        ("max_drawdown", _pct_ratio_to_percent(port_kpi.get("max_drawdown"))),
    ]:
        ws_p.append([k, v])

    ws_pd = wb.create_sheet("Portfolio_Daily")
    ws_pd.append(["Date", "Equity", "Invested", "GlobalCash", "CashAllocated", "PositionsValue", "Drawdown (%)"])
    ws_pd.freeze_panes = "A2"
    for cell in ws_pd[1]:
        cell.font = Font(bold=True)
    for r in port_daily:
        ws_pd.append([
            r.get("date"),
            _to_float(r.get("equity")),
            _to_float(r.get("invested")),
            _to_float(r.get("global_cash")),
            _to_float(r.get("cash_allocated")),
            _to_float(r.get("positions_value")),
            _pct_ratio_to_percent(r.get("drawdown")),
        ])
    _auto_width(ws_p)
    _auto_width(ws_pd)

    # -------- Charts (PNG for Sheets) --------
    if charts_enabled:
        ws_c = wb.create_sheet("Charts")
        ws_c.append(["Charts (images) – affichage compatible Google Sheets"])  # simple title
        ws_c["A1"].font = Font(bold=True)
        ws_c.append([
            "Note: pour rester léger, l’export compact limite le nombre de graphiques (chart_limit). "
            "Utilise ?chart_mode=all&chart_limit=XX si besoin."
        ])

        # Build list of (ticker, line_index, strat) candidates
        candidates = []
        for t, tentry in tickers_map.items():
            lines = (tentry or {}).get("lines") or []
            for strat in lines:
                ln = int(strat.get("line_index", 0) or 0) or 1
                fin = strat.get("final") or {}
                bt_final = _to_float(fin.get("BT"))
                candidates.append((t, ln, strat, bt_final if bt_final is not None else -10**9))

        # Decide which ones to chart
        selected = []
        if chart_mode == "first":
            # Use explicit ticker/line if possible, else fallback to first
            chosen = None
            if chart_ticker:
                for t, ln, strat, _score in candidates:
                    if t == chart_ticker and ln == chart_line:
                        chosen = (t, ln, strat)
                        break
            if chosen is None and candidates:
                t, ln, strat, _score = candidates[0]
                chosen = (t, ln, strat)
            if chosen is not None:
                selected = [chosen]
        else:
            # top or all
            candidates_sorted = sorted(candidates, key=lambda x: x[3], reverse=True)
            if chart_mode == "all":
                selected = [(t, ln, strat) for (t, ln, strat, _score) in candidates_sorted[: max(1, chart_limit)]]
                if len(candidates_sorted) > chart_limit:
                    ws_c.append([f"(Charts tronqués: {chart_limit} / {len(candidates_sorted)} couples (ticker, ligne).)"])
            else:
                # default: top
                selected = [(t, ln, strat) for (t, ln, strat, _score) in candidates_sorted[: max(1, chart_limit)]]

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from datetime import datetime
            from openpyxl.drawing.image import Image as XLImage
            from io import BytesIO

            def _parse_dates(date_strs):
                out = []
                for ds in date_strs:
                    if not ds:
                        out.append(None)
                        continue
                    try:
                        # expected ISO date
                        out.append(datetime.strptime(ds, "%Y-%m-%d"))
                    except Exception:
                        out.append(None)
                return out

            def _format_date_axis(ax):
                locator = mdates.AutoDateLocator(minticks=3, maxticks=9)
                ax.xaxis.set_major_locator(locator)
                ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))
                for label in ax.get_xticklabels():
                    label.set_rotation(0)
                    label.set_fontsize(8)

            def _plot_combo_bytes(ticker, ln, x_dates, y_sgn, y_bt, y_bmj, y_bmd, y_price, in_pos):
                """One image per (ticker,line) with 5 stacked subcharts."""
                buf = BytesIO()
                fig, axes = plt.subplots(5, 1, figsize=(10.5, 10.8), sharex=True)

                axes[0].plot(x_dates, y_sgn)
                axes[0].set_title(f"{ticker} L{ln} – S_G_N (%)")
                axes[0].set_ylabel("%")

                axes[1].plot(x_dates, y_bt)
                axes[1].set_title(f"{ticker} L{ln} – BT (%)")
                axes[1].set_ylabel("%")

                axes[2].plot(x_dates, y_bmj)
                axes[2].set_title(f"{ticker} L{ln} – BMJ (%)")
                axes[2].set_ylabel("%")

                axes[3].plot(x_dates, y_bmd)
                axes[3].set_title(f"{ticker} L{ln} – BMD (%)")
                axes[3].set_ylabel("%")

                # Price + position (stacked area style)
                # The user-facing rule is:
                #   - green area when we are in a BUY position (shares > 0)
                #   - red area when we are OUT (no BUY position)
                # We implement this by building two exclusive series and rendering a stacked area.
                price_in = [(y_price[i] if in_pos[i] else 0) for i in range(len(y_price))]
                price_out = [(y_price[i] if not in_pos[i] else 0) for i in range(len(y_price))]
                axes[4].stackplot(x_dates, price_in, price_out, alpha=0.55, colors=["green", "red"])
                axes[4].plot(x_dates, y_price, linewidth=1)
                axes[4].set_title(f"{ticker} L{ln} – Price + Position")
                axes[4].set_ylabel("Price")
                axes[4].set_xlabel("Date")
                _format_date_axis(axes[4])

                fig.tight_layout()
                fig.savefig(buf, format="png", dpi=140)
                plt.close(fig)
                buf.seek(0)
                return buf

            def _plot_portfolio_bytes(x_dates, equity, dd_pct):
                """Portfolio image with equity curve + drawdown."""
                buf = BytesIO()
                fig, axes = plt.subplots(2, 1, figsize=(10.5, 5.4), sharex=True)
                axes[0].plot(x_dates, equity)
                axes[0].set_title("Portfolio – Equity")
                axes[0].set_ylabel("Equity")

                axes[1].plot(x_dates, dd_pct)
                axes[1].set_title("Portfolio – Drawdown (%)")
                axes[1].set_ylabel("%")
                axes[1].set_xlabel("Date")
                _format_date_axis(axes[1])

                fig.tight_layout()
                fig.savefig(buf, format="png", dpi=140)
                plt.close(fig)
                buf.seek(0)
                return buf

            # Insert images stacked vertically
            anchor_row = 4
            if port_daily:
                x_dates = _parse_dates([r.get("date") for r in port_daily])
                equity = [_to_float(r.get("equity")) for r in port_daily]
                dd_pct = [_pct_ratio_to_percent(r.get("drawdown")) for r in port_daily]
                img = XLImage(_plot_portfolio_bytes(x_dates, equity, dd_pct))
                ws_c.add_image(img, f"A{anchor_row}")
                anchor_row += 26
            for (t, ln, strat) in selected:
                daily = strat.get("daily") or []
                x_raw = [r.get("date") for r in daily]
                x_dates = _parse_dates(x_raw)
                y_sgn = [_pct(r.get("S_G_N")) for r in daily]
                y_bt = [_pct(r.get("BT")) for r in daily]
                y_bmj = [_pct(r.get("BMJ")) for r in daily]
                y_bmd = [_pct(r.get("BMD")) for r in daily]
                y_price = [_to_float(r.get("price_close")) for r in daily]
                in_pos = [(_to_float(r.get("shares")) or 0) > 0 for r in daily]

                img = XLImage(_plot_combo_bytes(t, ln, x_dates, y_sgn, y_bt, y_bmj, y_bmd, y_price, in_pos))
                ws_c.add_image(img, f"A{anchor_row}")
                anchor_row += 32  # spacing between blocks

        except Exception:
            # If matplotlib isn't available or something fails, keep workbook without charts.
            ws_c.append(["(Charts indisponibles dans cet environnement)"])

    # Return file
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="backtest_{bt.id}_export_compact.xlsx"'
    return response


def _safe_fs_segment(value: str) -> str:
    """Filesystem-safe segment (must match parquet_storage._safe_segment semantics)."""
    import re

    value = (value or "").strip()
    if not value:
        return "unknown"
    value = value.lower()
    value = re.sub(r"[^a-z0-9_-]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "unknown"


def _resolve_backtest_parquet_dir(bt: Backtest) -> Path:
    """Return the expected parquet directory for a backtest.

    Layout:
        <BACKTEST_DATA_DIR>/backtests/<id>/<scenario_segment>/
    """
    base_dir = os.environ.get("BACKTEST_DATA_DIR", "/data").strip() or "/data"
    scenario_name = getattr(getattr(bt, "scenario", None), "name", "")
    scenario_segment = _safe_fs_segment(scenario_name) if scenario_name else str(bt.scenario_id or "scenario")
    return Path(base_dir) / "backtests" / str(bt.id) / scenario_segment


@login_required

def _arrow_table_to_csv_safe(table):
    """Convert Arrow table to CSV-writable by stringifying nested types (list/struct/map).
    This is used only for the optional details CSV export to avoid pyarrow errors.
    """
    try:
        import pyarrow as pa  # type: ignore
    except Exception:
        return table

    cols = []
    names = []
    for name in table.column_names:
        col = table[name]
        t = col.type
        # Nested / complex types are not supported by pyarrow.csv.write_csv
        if pa.types.is_list(t) or pa.types.is_large_list(t) or pa.types.is_struct(t) or pa.types.is_map(t):
            # Convert each cell to compact JSON string (preserves information).
            pylist = col.to_pylist()
            strlist = [json.dumps(v, ensure_ascii=False) if v is not None else None for v in pylist]
            col = pa.array(strlist, type=pa.string())
        cols.append(col)
        names.append(name)
    return pa.table(cols, names=names)

def backtest_export_details(request, pk: int):
    """Export backtest daily series as a ZIP of Parquet files (or CSV).

    This is an additive export and does not modify the legacy Excel exports.
    """
    bt = get_object_or_404(Backtest, pk=pk)

    fmt = (request.GET.get("format") or "parquet").strip().lower()
    if fmt not in {"parquet", "csv"}:
        fmt = "parquet"

    # For safety: details export is only meaningful when Parquet storage exists.
    parquet_dir = _resolve_backtest_parquet_dir(bt)
    if not parquet_dir.exists() or not parquet_dir.is_dir():
        messages.error(
            request,
            "Export détails indisponible : fichiers Parquet non trouvés. "
            "Relance le backtest avec ENABLE_PARQUET_STORAGE=1 (et un volume /data persistant).",
        )
        return redirect("backtest_detail", pk=bt.id)

    # Collect parquet files
    parquet_files = sorted([p for p in parquet_dir.glob("*.parquet") if p.is_file()])
    if not parquet_files:
        messages.error(
            request,
            "Export détails indisponible : aucun fichier Parquet dans le dossier de ce backtest.",
        )
        return redirect("backtest_detail", pk=bt.id)

    # Create ZIP on disk to avoid memory blowups
    tmp = tempfile.NamedTemporaryFile(prefix=f"backtest_{bt.id}_details_", suffix=".zip", delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()

    try:
        with pyzip.ZipFile(tmp_path, "w", compression=pyzip.ZIP_DEFLATED) as zf:
            if fmt == "parquet":
                for fp in parquet_files:
                    zf.write(fp, arcname=fp.name)
            else:
                # CSV export requires pyarrow to read parquet (optional dependency)
                try:
                    import pyarrow.parquet as pq  # type: ignore
                    import pyarrow.csv as pacsv  # type: ignore
                except Exception as e:
                    messages.error(request, f"pyarrow requis pour l'export CSV: {e}")
                    return redirect("backtest_detail", pk=bt.id)

                for fp in parquet_files:
                    table = pq.read_table(fp)
                    table = _arrow_table_to_csv_safe(table)
                    buf = BytesIO()
                    pacsv.write_csv(table, buf)
                    buf.seek(0)
                    zf.writestr(fp.with_suffix(".csv").name, buf.getvalue())

        tmp_f = open(tmp_path, "rb")
        filename = f"backtest_{bt.id}_details_{fmt}.zip"
        resp = FileResponse(tmp_f, as_attachment=True, filename=filename, content_type="application/zip")
        return resp
    finally:
        # FileResponse will keep the file handle open for streaming.
        # We keep the temp file on disk; OS cleanup can be handled by a cron, or you can
        # switch to a managed temp dir in Vultr if needed.
        pass


@login_required
def indicators_help(request):
    """UI helper: explain how indicators and alerts are computed.

    This page is intentionally static (documentation only) to avoid any risk of breaking business rules.
    """
    return render(request, "help_indicators.html")