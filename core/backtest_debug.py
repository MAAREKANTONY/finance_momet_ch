from __future__ import annotations

from typing import Any, Iterable
from openpyxl import Workbook

from .excel_utils import append_excel_row, normalize_excel_cell
from .models import Backtest


PREFERRED_DATA_COLUMNS = [
    'date', 'open', 'high', 'low', 'close',
    'P', 'P_t', 'P_STUDY', 'DELTA', 'DELTA_J',
    'M', 'X', 'M1', 'X1', 'T', 'Q', 'S',
    'K1', 'K2', 'K3', 'K4', 'KF', 'K1F', 'K2F', 'KF3',
    'SUM_SLOPE', 'SLOPE_VRAI', 'GM', 'GM_POS', 'GM_NEG', 'GM_NEU',
    'buy_signal', 'sell_signal', 'action', 'in_position',
    'NB_JOUR_OUVRES', 'BUY_DAYS_CLOSED',
    'TRADABLE_DAYS', 'TRADABLE_DAYS_NOT_IN_POSITION', 'TRADABLE_DAYS_IN_POSITION_CLOSED',
    'RATIO_NOT_IN_POSITION', 'RATIO_IN_POSITION',
]


FORMULA_ROWS = [
    ('P(t)', 'Prix d’étude', '(a*Close + b*High + c*Low + d*Open) / (a+b+c+d)', '=(a*Close+b*High+c*Low+d*Open)/(a+b+c+d)'),
    ('δj(t)', 'Variation journalière du prix d’étude', '(P(t)-P(t-1))/P(t-1)', '=(P_t-P_t_1)/P_t_1'),
    ('M(t)', 'Maximum glissant sur N1', 'max(P(t-N1+1..t))', '=MAX(plage_P_N1)'),
    ('X(t)', 'Minimum glissant sur N1', 'min(P(t-N1+1..t))', '=MIN(plage_P_N1)'),
    ('M1/X1', 'Lissages / extrêmes secondaires', 'Selon implémentation moteur', 'Reprendre avec fenêtres du scénario'),
    ('T/Q/S', 'Amplitudes / bandes', 'Selon implémentation moteur', 'Reprendre depuis les colonnes exportées'),
    ('K1..K4', 'Lignes fixes', 'Selon implémentation moteur', 'Comparer directement aux colonnes exportées'),
    ('Kf / K1f / K2f / Kf3', 'Lignes flottantes', 'Selon implémentation moteur', 'Comparer directement aux colonnes exportées'),
    ('SUM_SLOPE', 'Pente cumulée', 'Somme glissante des variations retenues', '=SUM(plage_variations)'),
    ('SLOPE_VRAI', 'Pente vraie', '(P(t)-P(t-Npente))/P(t-Npente)', '=(P_t-P_t_Npente)/P_t_Npente'),
    ('GM', 'Global momentum', 'Moyenne des performances sur Nglobal', '=AVERAGE(plage_delta_Nglobal)'),
    ('GM_POS/NEG/NEU', 'Discrétisation GM', 'Selon seuils moteur', 'Comparer la valeur GM au seuil'),
    ('BUY / SELL', 'Décisions de trading', 'Conditions de ligne + éventuel filtre GM', 'Vérifier les colonnes de signaux puis action'),
    ('BT', 'Performance backtest', 'Selon métrique moteur', 'Comparer à la ligne finale'),
    ('BMD', 'Performance moyenne par jour tradable', 'Selon métrique moteur', 'Comparer à la ligne finale'),
]


def _codes_to_label(v: Any) -> str:
    if isinstance(v, (list, tuple)):
        return ' + '.join(str(x) for x in v if str(x).strip())
    if v is None:
        return ''
    return str(v)


def _int_or_zero(v: Any) -> int:
    try:
        return int(v)
    except Exception:
        return 0


def augment_debug_row(row: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(row, dict):
        return row
    d = dict(row)
    nip = _int_or_zero(d.get('NB_JOUR_OUVRES'))
    ipc = _int_or_zero(d.get('BUY_DAYS_CLOSED'))
    td = nip + ipc
    d['TRADABLE_DAYS_NOT_IN_POSITION'] = nip
    d['TRADABLE_DAYS_IN_POSITION_CLOSED'] = ipc
    d['TRADABLE_DAYS'] = td
    d['RATIO_NOT_IN_POSITION'] = (nip / td * 100.0) if td > 0 else 0.0
    d['RATIO_IN_POSITION'] = (ipc / td * 100.0) if td > 0 else 0.0
    return d


def _load_daily_from_line(line: dict[str, Any]) -> list[dict[str, Any]]:
    try:
        from .services.backtesting.results_offload import load_daily_from_line
        daily = load_daily_from_line(line or {})
    except Exception:
        daily = (line or {}).get('daily') or []
    return [augment_debug_row(dict(r)) for r in (daily or []) if isinstance(r, dict)]


def get_backtest_debug_payload(bt: Backtest, ticker: str = '', line: str | int | None = None) -> dict[str, Any]:
    results = bt.results or {}
    tickers_map = results.get('tickers') or {}
    if not tickers_map:
        raise ValueError("Aucun résultat disponible pour ce backtest.")

    selected_ticker = (ticker or '').strip() or next(iter(tickers_map.keys()))
    if selected_ticker not in tickers_map:
        raise ValueError("Ticker introuvable dans les résultats du backtest.")
    tentry = tickers_map.get(selected_ticker) or {}
    lines = tentry.get('lines') or []
    if not lines:
        raise ValueError("Aucune ligne disponible pour ce ticker.")

    selected_line_index = None
    if line not in (None, ''):
        try:
            selected_line_index = int(line)
        except Exception:
            selected_line_index = None

    selected_line = None
    if selected_line_index is not None:
        selected_line = next((ln for ln in lines if int(ln.get('line_index') or 0) == selected_line_index), None)
    if selected_line is None:
        selected_line = lines[0]
        selected_line_index = int(selected_line.get('line_index') or 1)

    daily = _load_daily_from_line(selected_line)
    final = augment_debug_row(dict((selected_line.get('final') or {})))
    return {
        'ticker': selected_ticker,
        'line_index': selected_line_index,
        'line': selected_line,
        'daily': daily,
        'final': final,
        'buy_label': _codes_to_label(selected_line.get('buy')),
        'sell_label': _codes_to_label(selected_line.get('sell')),
        'results_meta': results.get('meta') or {},
    }


def _ordered_columns(rows: Iterable[dict[str, Any]]) -> list[str]:
    seen = []
    seen_set = set()
    for c in PREFERRED_DATA_COLUMNS:
        if c not in seen_set:
            seen.append(c)
            seen_set.add(c)
    for row in rows:
        for k in row.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    return [c for c in seen if any(c in row for row in rows)]


def _append_kv(ws, key: str, value: Any) -> None:
    append_excel_row(ws, [key, normalize_excel_cell(value)])


def build_backtest_debug_workbook(bt: Backtest, ticker: str = '', line: str | int | None = None):
    payload = get_backtest_debug_payload(bt, ticker=ticker, line=line)
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = 'DATA'

    daily = payload['daily'] or []
    columns = _ordered_columns(daily)
    if columns:
        append_excel_row(ws_data, columns)
        for row in daily:
            append_excel_row(ws_data, [row.get(c, '') for c in columns])
    else:
        append_excel_row(ws_data, ['info'])
        append_excel_row(ws_data, ['Aucune ligne quotidienne disponible'])

    ws_formulas = wb.create_sheet('FORMULAS')
    append_excel_row(ws_formulas, ['Section', 'Valeur'])
    _append_kv(ws_formulas, 'Backtest ID', bt.id)
    _append_kv(ws_formulas, 'Backtest name', bt.name)
    _append_kv(ws_formulas, 'Scenario', getattr(bt.scenario, 'name', ''))
    _append_kv(ws_formulas, 'Ticker', payload['ticker'])
    _append_kv(ws_formulas, 'Line index', payload['line_index'])
    _append_kv(ws_formulas, 'BUY conditions', payload['buy_label'])
    _append_kv(ws_formulas, 'SELL conditions', payload['sell_label'])
    _append_kv(ws_formulas, 'Start date', bt.start_date)
    _append_kv(ws_formulas, 'End date', bt.end_date)
    _append_kv(ws_formulas, 'Capital total', bt.capital_total)
    _append_kv(ws_formulas, 'Capital per ticker', bt.capital_per_ticker)
    _append_kv(ws_formulas, 'Capital mode', bt.capital_mode)
    _append_kv(ws_formulas, 'Ratio threshold', bt.ratio_threshold)
    _append_kv(ws_formulas, 'Include all tickers', bt.include_all_tickers)
    _append_kv(ws_formulas, 'Warmup days', bt.warmup_days)
    _append_kv(ws_formulas, 'Close positions at end', bt.close_positions_at_end)

    scenario = bt.scenario
    append_excel_row(ws_formulas, [])
    append_excel_row(ws_formulas, ['Scenario parameter', 'Value'])
    for field in [
        'a', 'b', 'c', 'd', 'e', 'vc', 'fl', 'n1', 'n2', 'n3', 'n4', 'n5', 'k2j', 'cr',
        'n5f3', 'crf3', 'npente', 'nglobal', 'slope_threshold', 'npente_basse', 'slope_threshold_basse',
    ]:
        if hasattr(scenario, field):
            _append_kv(ws_formulas, field, getattr(scenario, field))

    append_excel_row(ws_formulas, [])
    append_excel_row(ws_formulas, ['Formula', 'Meaning', 'Theory', 'Excel hint'])
    for row in FORMULA_ROWS:
        append_excel_row(ws_formulas, list(row))

    ws_final = wb.create_sheet('FINAL')
    append_excel_row(ws_final, ['Metric', 'Value'])
    for k, v in payload['final'].items():
        append_excel_row(ws_final, [k, v])

    for ws in [ws_data, ws_formulas, ws_final]:
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells[:200]) if hasattr(column_cells, '__iter__') else 10
            letter = column_cells[0].column_letter
            ws.column_dimensions[letter].width = min(max(length + 2, 12), 40)

    filename = f"backtest_{bt.id}_{payload['ticker']}_line_{payload['line_index']}_debug.xlsx"
    return wb, filename
